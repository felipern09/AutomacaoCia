import shutil
from datetime import datetime as dt, timedelta as td
import datetime
from dateutil.relativedelta import relativedelta
from email import encoders
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
import holidays
import hashlib
import locale
from openpyxl import load_workbook as l_w
from openpyxl.styles import PatternFill, Font
import openpyxl.utils.cell
import os
import pandas as pd
from PIL import ImageGrab
import pyautogui as pa
from src.models.modelsfolha import Aula, Folha, Aulas, Faltas, Ferias, Hrcomplement, Atestado, Desligados, \
    Escala, Substituicao, enginefolha
import smtplib
from src.models.dados_servd import em_rem, em_ti, em_if, k1, host, port, rede, em_fin, em_lgpd, pasta_dexion, \
    pasta_estag, pasta_func
from src.models.models import Colaborador, engine
from sqlalchemy.orm import sessionmaker,declarative_base
from sqlalchemy import create_engine, MetaData, Column, Integer, String
import zipfile
import tkinter.filedialog
from tkinter import messagebox
import tkinter.filedialog
import time as t
import win32com.client as client
from pdf2image import convert_from_path
from PIL import Image, ImageDraw, ImageFont, ImageOps
from fpdf import FPDF

locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')

atestado = PatternFill(start_color='A9D08E',
                       end_color='A9D08E',
                       fill_type='solid')
falta = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')
ferias = PatternFill(start_color='9BC2E6',
                     end_color='9BC2E6',
                     fill_type='solid')
feriado = PatternFill(start_color='F4B084',
                      end_color='F4B084',
                      fill_type='solid')
fds = PatternFill(start_color='BFBFBF',
                  end_color='BFBFBF',
                  fill_type='solid')
deslig = PatternFill(start_color='454545',
                     end_color='454545',
                     fill_type='solid')
subst = PatternFill(start_color='FFFF00',
                    end_color='FFFF00',
                    fill_type='solid')
comple = PatternFill(start_color='FFC000',
                     end_color='FFC000',
                     fill_type='solid')


def confirma_folha(comp: int):
    """
    Confirm that the user wnats to proceed with procedures to post values of payroll in external aplication.
    :param comp: Month reference to payroll calculate.
    :return: Call function lancar_folha_no_dexion().
    """
    resp = messagebox.askyesno(title='Tem certeza?',
                               message=f'Tem certeza que deseja lançar a folha do mês {comp} no Dexion?')
    if resp:
        lancar_folha_no_dexion(comp)


def confirma_grade(comp: str):
    """
    Confirm that the user wnats to proceed with procedures to register the payroll.
    :param comp: Month reference to payroll calculate.
    :return: Call function salvar_planilha_soma_final().
    """
    dia, mes, ano = comp.split('/')
    mes = int(mes)
    ano = int(ano)
    r = messagebox.askyesno(title='Tem certeza?',
                            message=f'Tem certeza que deseja gerar a folha do mês {mes}/{ano}?\n'
                                    f'Essa ação irá sobrepor qualquer arquivo de folha já salvo na pasta dessa competência.')
    if r:
        salvar_planilha_soma_final(mes, ano)


def lancar_folha_no_dexion(competencia):
    """
    Register values of payroll in external aplication through PyAutogui package.
    :param competencia: Month reference to payroll calculate.
    :return: External aplication with payroll values registred.
    """
    while 1 < 2:
        if pa.locateOnScreen('../models/static/imgs/Dexion.png'):
            pa.click(pa.center(pa.locateOnScreen('../models/static/imgs/Dexion.png')))
            break
        else:
            t.sleep(5)
    pa.press('a'), t.sleep(2)
    hj = dt.today()
    mes = str(competencia).zfill(2)
    ano = hj.year
    mesext = {'01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR', '05': 'MAI', '06': 'JUN',
              '07': 'JUL', '08': 'AGO', '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'}
    folhagrd = os.path.abspath(rf"\\192.168.0.250\rh\01 - RH\01 - Administração.Controles\04 - Folha de Pgto\{ano}\{mes} - {mesext[mes]}\Grades e Comissões\Lancamentos.xlsx")
    wb = l_w(folhagrd, read_only=False)

    # lançamento de faltas
    sh = wb['Faltas']
    x = 2
    while x <= len(sh['A']):
        mat = str(sh[f'A{x}'].value)
        rub = str(sh[f'C{x}'].value)
        hr = str(sh[f'D{x}'].value)
        pa.write(mat), t.sleep(3), pa.press('enter', 2), t.sleep(2.5), pa.press('i'), t.sleep(2.5), pa.write(rub)
        t.sleep(1), pa.press('enter'), t.sleep(1.5), pa.write(hr), t.sleep(1.5), pa.press('enter', 65), t.sleep(4)
        x += 1

    # deletar férias antigas
    sh = wb['DeletarFerias']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            rub = ['1006', '1007', '1010', '1011', '1012', '1037']
            mat = str(sh[f'A{x}'].value)
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(1), pa.press('d')
            for r in rub:
                pa.write(r), t.sleep(0.8), pa.press('enter'), t.sleep(0.5), pa.press('left'), t.sleep(0.5), pa.press(
                    'enter')
            pa.press('enter', 2)
            x += 1

    # lançamento de horistas
    sh = wb['Horistas']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            mat = str(sh[f'A{x}'].value)
            rub = str(sh[f'C{x}'].value)
            hr = str(sh[f'D{x}'].value)
            obshr = str(sh[f'E{x}'].value)
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(2.3)
            pa.press('a'), t.sleep(0.5), pa.write(rub)
            pa.press('enter'), t.sleep(0.5), pa.write(hr), t.sleep(0.5), pa.press('enter', 2)
            pa.press('enter')
            x += 1
    # lançamento de dias dobrados — estágio
    sh = wb['Compl Est']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            mat = str(sh[f'A{x}'].value)
            rub = str(sh[f'C{x}'].value)
            dias = str(sh[f'D{x}'].value)
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(0.5), pa.press('i'), t.sleep(0.5), pa.write(rub)
            pa.press('enter'), t.sleep(0.5), pa.write(dias), t.sleep(0.5), pa.press('enter')
            pa.press('enter'), t.sleep(0.5), pa.press('enter')
            x += 1

    # lançamento de comissões
    sh = wb['Comissoes']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            mat = str(sh[f'A{x}'].value)
            rub = str(sh[f'C{x}'].value)
            hr = str(sh[f'D{x}'].value)
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(0.5), pa.press('i'), t.sleep(0.5), pa.write(rub)
            pa.press('enter'), t.sleep(0.5), pa.write(hr), t.sleep(0.5), pa.press('enter')
            pa.press('enter'), t.sleep(0.5), pa.press('enter')
            x += 1

    # # Lançamento de adiantamento
    sh = wb['Adiantamento']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            mat = str(sh[f'A{x}'].value)
            rub = '81'
            hr = str(sh[f'D{x}'].value)
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(0.5), pa.press('i'), t.sleep(0.5), pa.write(rub)
            pa.press('enter'), t.sleep(0.5), pa.write(hr), t.sleep(0.5), pa.press('enter')
            pa.press('enter'), t.sleep(0.5), pa.press('enter')
            x += 1

    # lançamento de desconto de VT
    sh = wb['DescontoVT']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            mat = str(sh[f'A{x}'].value)
            rub = '80'
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(0.5), pa.press('i'), t.sleep(0.5), pa.write(rub)
            pa.press('enter'), t.sleep(0.5), pa.press('enter')
            pa.press('enter'), t.sleep(0.5), pa.press('enter')
            x += 1

    # lançamento de plano de saúde
    sh = wb['Plano']
    x = 2
    while x <= len(sh['A']):
        if sh[f'A{x}'].value is not None:
            mat = str(sh[f'A{x}'].value)
            rub = str(sh[f'C{x}'].value)
            hr = str(sh[f'D{x}'].value)
            sq = str(sh[f'E{x}'].value)
            pa.write(mat), t.sleep(0.5), pa.press('enter', 2), t.sleep(0.5), pa.press('i'), t.sleep(0.5), pa.write(rub)
            pa.press('enter'), t.sleep(0.5), pa.write(sq), t.sleep(0.5), pa.press('enter'), t.sleep(0.5), pa.write(hr)
            pa.press('enter', 3), t.sleep(0.5),
            x += 1
    while 1 < 2:
        if pa.locateOnScreen('../models/static/imgs/pyt.png'):
            pa.click(pa.center(pa.locateOnScreen('../models/static/imgs/pyt.png')))
            break
        else:
            t.sleep(5)
    tkinter.messagebox.showinfo(
        title='Folha ok!',
        message=f'Folha do mês {competencia} lançada no Dexion com sucesso!'
    )


def somar_aulas_da_grade(diasem: str, inic: datetime.datetime, fim: datetime.datetime, competencia: int, iniciograd: str, fimgrad: str) -> float:
    """
    Sum classes that have status 'Active' at the period selected for the payroll.
    :param diasem: Weekday.
    :param inic: Class start time.
    :param fim: End time of the class.
    :param competencia: Month of the payroll.
    :param iniciograd: Start day for the payroll.
    :param fimgrad: End day of the payroll.
    :return: Sum of all classes times for each day of the week.
    """
    horas = fim - inic
    hr, minut, seg = str(horas).split(':')
    igrad = dt.strptime(iniciograd, '%d/%m/%Y')
    if fimgrad is not None:
        fgrad = dt.strptime(fimgrad, '%d/%m/%Y')
    else:
        fgrad = ''
    somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
    somadia = round(somadia, 2)
    soma = 0
    comp = dt(day=1, month=competencia, year=dt.today().year)
    inicio = dt(day=21, month=(comp - relativedelta(months=1)).month, year=(comp - relativedelta(months=1)).year)
    fechamento = dt(day=20, month=comp.month, year=comp.year)

    def intervalo(inicio, fechamento):
        for n in range(int((fechamento - inicio).days) + 1):
            yield inicio + td(n)

    if fgrad != '':
        if diasem == 'Segunda':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 0 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Terça':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 1 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Quarta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 2 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Quinta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 3 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Sexta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 4 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Sábado':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 5 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Domingo':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 6 and igrad <= dia <= fgrad:
                    soma += somadia
    else:
        if diasem == 'Segunda':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 0 and dia >= igrad:
                    soma += somadia
        if diasem == 'Terça':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 1 and dia >= igrad:
                    soma += somadia
        if diasem == 'Quarta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 2 and dia >= igrad:
                    soma += somadia
        if diasem == 'Quinta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 3 and dia >= igrad:
                    soma += somadia
        if diasem == 'Sexta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 4 and dia >= igrad:
                    soma += somadia
        if diasem == 'Sábado':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 5 and dia >= igrad:
                    soma += somadia
        if diasem == 'Domingo':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 6 and dia >= igrad:
                    soma += somadia
    return round(soma, 2)


def listar_aulas_ativas(compet) -> list:
    """
    List all classes with 'Active' status.
    :return: List of active classes.
    """
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    competencia = dt(day=10, month=compet, year=dt.today().year)
    inicio = dt(day=21, month=(competencia - relativedelta(months=1)).month,
                year=(competencia - relativedelta(months=1)).year)

    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    for aulat in aulasativasdb:
        pessoa = sessioncol.query(Colaborador).filter_by(nome=aulat.professor).order_by(Colaborador.matricula.desc()).first()
        if pessoa:
            if pessoa.desligamento is not None:
                if dt.strptime(pessoa.desligamento, '%d/%m/%Y') >= inicio:
                    pass
                else:
                    aulat.status = 'Inativa'
                    session.commit()

    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    aula = []
    for i, a in enumerate(aulasativasdb):
        aula.append(i)
        aula[i] = Aula(a.nome, a.professor, a.departamento, a.diadasemana, a.inicio, a.fim, a.valor, a.iniciograde,
                       a.fimgrade)
        yield aula[i]
    session.close()
    return aula


def listar_departamentos_ativos() -> list:
    """
    List all departments with classes with "Active" status.
    :return: List of active departments.
    """
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    departamentos = []
    for i, a in enumerate(aulasativasdb):
        departamentos.append(a.departamento)
        departamentos = list(set(departamentos))
    session.close()
    return departamentos


def listar_professores_ativos() -> list:
    """
    List of all teachers who have active classes.
    :return: List of active teachers.
    """
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    professores = []
    for i, a in enumerate(aulasativasdb):
        professores.append(a.professor)
        professores = list(set(professores))
    session.close()
    return professores


def calcular_total_monetario_folha(compet: int) -> float:
    """
    Sum the total payroll payment.
    :param compet: Month of payroll
    :return: Payment amount.
    """
    somatorio = 0
    for al in listar_aulas_ativas(compet):
        somatorio += somar_aulas_da_grade(al.dia, al.inicio, al.fim, compet, al.iniciograde, al.fimgrade) * float(
            str(al.valor).replace(',', '.')) * al.dsr
    return round(somatorio, 2)


def somar_horas_professor(folha: Folha, prof: str, depto: str, nome: str, compet: int) -> float:
    """
    Sum total of hours for each teacher.
    :param folha: Payroll reference.
    :param prof: Teacher.
    :param depto: Department
    :param nome: Name of the class.
    :param compet: Month of payroll
    :return: The amount of hours for each teacher on this payroll.
    """
    somahoras = 0
    for aula in folha.aulas:
        if aula.professor == prof and aula.departamento == depto and aula.nome == nome:
            somahoras += somar_aulas_da_grade(aula.dia, aula.inicio, aula.fim, compet, aula.iniciograde, aula.fimgrade)
    return somahoras


def consultar_faltas(comp, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    falt = session.query(Faltas).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for f in falt:
        if inicio <= dt.strptime(f.data, '%d/%m/%Y') <= fim:
            if f.professor in dic:
                d2 = {f.professor: {f.data: {f.departamento: f.horas}}}
                dic[f.professor] = {**dic[f.professor], **d2[f.professor]}
            else:
                d2 = {f.professor: {f.data: {f.departamento: f.horas}}}
                dic = {**dic, **d2}
    session.close()
    return dic


def consultar_ferias(comp, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    fer = session.query(Ferias).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for f in fer:
        if inicio <= dt.strptime(f.inicio, '%d/%m/%Y') <= fim:
            if f.professor in dic:
                d2 = {f.professor: {f.departamento: {f.inicio: f.fim}}}
                dic[f.professor] = {**dic[f.professor], **d2[f.professor]}
            else:
                d2 = {f.professor: {f.departamento: {f.inicio: f.fim}}}
                dic = {**dic, **d2}
        if inicio <= dt.strptime(f.fim, '%d/%m/%Y') <= fim:
            if f.professor in dic:
                d2 = {f.professor: {f.departamento: {f.inicio: f.fim}}}
                dic[f.professor] = {**dic[f.professor], **d2[f.professor]}
            else:
                d2 = {f.professor: {f.departamento: {f.inicio: f.fim}}}
                dic = {**dic, **d2}
    session.close()
    return dic


def consultar_atestados(comp, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    ates = session.query(Atestado).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for a in ates:
        if inicio <= dt.strptime(a.data, '%d/%m/%Y') <= fim:
            if a.professor in dic:
                d2 = {a.professor: {a.data: a.departamento}}
                dic[a.professor] = {**dic[a.professor], **d2[a.professor]}
            else:
                d2 = {a.professor: {a.data: a.departamento}}
                dic = {**dic, **d2}
    session.close()
    return dic


def listar_feriados(comp: int, ano) -> list:
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    # Get the Bank Holidays for the given country
    feriados = holidays.country_holidays('BR')
    # Create a list of dates between the start and end date
    intervalo_datas = pd.date_range(inicio, fim)
    # Filter the dates to only include Bank Holidays
    feriados_nacionais = [date for date in intervalo_datas if date in feriados]
    return feriados_nacionais


def consultar_substituicoes(comp: int, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    subst = session.query(Substituicao).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for s in subst:
        if inicio <= dt.strptime(s.data, '%d/%m/%Y') <= fim:
            dic.update({s.numero: {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}})
    session.close()
    return dic


def consultar_desligamentos(comp: int, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    desl = session.query(Desligados).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for d in desl:
        if inicio <= dt.strptime(d.datadesligamento, '%d/%m/%Y') <= fim:
            if d.professor in dic:
                d2 = {d.professor: {d.departamento: d.datadesligamento}}
                dic[d.professor] = {**dic[d.professor], **d2[d.professor]}
            else:
                d2 = {d.professor: {d.departamento: d.datadesligamento}}
                dic = {**dic, **d2}
    session.close()
    return dic


def consultar_escalas(comp: int, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    esc = session.query(Escala).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for e in esc:
        if inicio <= dt.strptime(e.data, '%d/%m/%Y') <= fim:
            if e.professor in dic:
                d2 = {e.professor: {e.data: {e.departamento: e.horas}}}
                dic[e.professor] = {**dic[e.professor], **d2[e.professor]}
            else:
                d2 = {e.professor: {e.data: {e.departamento: e.horas}}}
                dic = {**dic, **d2}
    session.close()
    return dic


def consultar_horas_complementares(comp: int, ano) -> dict:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    hrsc = session.query(Hrcomplement).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).month,
                year=(dt(day=1, month=comp, year=ano) - relativedelta(months=1)).year)
    fim = dt(day=20, month=comp, year=ano)
    dic = {}
    for h in hrsc:
        if inicio <= dt.strptime(h.data, '%d/%m/%Y') <= fim:
            if h.professor in dic:
                d2 = {h.professor: {h.data: {h.departamento: h.horas}}}
                dic[h.professor] = {**dic[h.professor], **d2[h.professor]}
            else:
                d2 = {h.professor: {h.data: {h.departamento: h.horas}}}
                dic = {**dic, **d2}
    session.close()
    return dic


def salvar_planilha_grade_horaria(dic: dict, comp: int, an: int):
    hj = dt.today()
    mes = str(comp).zfill(2)
    ano = an
    mesext = {'01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR', '05': 'MAI', '06': 'JUN',
              '07': 'JUL', '08': 'AGO', '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'}
    pasta_pgto = rf'\\192.168.0.250\rh\01 - RH\01 - Administração.Controles\04 - Folha de Pgto\{ano}\{mes} - {mesext[mes]}\Grades e Comissões'

    grd = os.path.relpath(rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\static\files\Grade.xlsx')
    grade = l_w(grd, read_only=False)
    plan1 = grade['Planilha1']
    flt = consultar_faltas(comp, an)
    subs = consultar_substituicoes(comp, an)
    dslg = consultar_desligamentos(comp, an)
    fer = consultar_ferias(comp, an)
    complem = consultar_horas_complementares(comp, an)
    atest = consultar_atestados(comp, an)
    feriad = listar_feriados(comp, an)
    escal = consultar_escalas(comp, an)
    competencia = dt(day=10, month=comp, year=an)
    inicio = dt(day=21, month=(competencia - relativedelta(months=1)).month,
                year=(competencia - relativedelta(months=1)).year)
    fechamento = dt(day=20, month=competencia.month, year=competencia.year)
    plan1['A1'].value = 'Folha'
    plan1['B1'].value = f'{fechamento.month} de {fechamento.year}'

    def intervalo(inicio, fechamento):
        for n in range(int((fechamento - inicio).days) + 1):
            yield dt.strftime(inicio + td(n), '%d/%m/%Y')

    def descobrir_ano(diadt: str):
        ddt, m = diadt.split('/')
        if int(ddt) > fechamento.day:
            anodt = inicio.year
        else:
            anodt = fechamento.year
        return anodt

    # inserir data no cabeçalho da grade
    col = 3
    for item in list(intervalo(inicio, fechamento)):
        plan1.cell(column=col, row=3, value=dt.strftime(dt.strptime(item, '%d/%m/%Y'), '%a'))
        plan1.cell(column=col, row=4, value=dt.strftime(dt.strptime(item, '%d/%m/%Y'), '%d/%m'))
        col += 1
    # insesir total na coluna ao final das datas
    plan1.cell(column=col, row=3, value='Total')

    # separar cada prof de cada depto
    musculacao = []
    ginastica = []
    esportes = []
    kids = []
    cross = []
    mensalistas = []
    pnt = []
    for i in dic:
        for sub in dic[i]:
            if dic[i][sub] == {}:
                pass
            else:
                if sub == 'Musculação':
                    musculacao.append(i)
                    musculacao.sort()
                if sub == 'Ginástica':
                    ginastica.append(i)
                    ginastica.sort()
                if sub == 'Esportes':
                    esportes.append(i)
                    esportes.sort()
                if sub == 'Kids':
                    kids.append(i)
                    kids.sort()
                if sub == 'Cross Cia':
                    cross.append(i)
                    cross.sort()
                if sub == 'PNT':
                    pnt.append(i)
                    pnt.sort()
                if sub == 'Mensalistas':
                    mensalistas.append(i)
                    mensalistas.sort()

    plan1['A5'].value = 'Musculação'
    novalinha = 6
    for i in musculacao:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Musculação'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Musculação' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                for numb in subs:
                    for nome in subs[numb]:
                        for substituto in subs[numb][nome]:
                            for depart in subs[numb][nome][substituto]:
                                for dia in subs[numb][nome][substituto][depart]:
                                    if depart == 'Musculação' and nome == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).fill = falta
                                    if depart == 'Musculação' and substituto == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                column=cell.column, row=novalinha).value + float(
                                                str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                            plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                # {d.professor: {d.departamento: d.datadesligamento}}
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Musculação' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                      month=int(str(plan1.cell(column=cell.column,
                                                                                               row=cell.row + 1).value).split(
                                                                              '/')[1]),
                                                                      year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                # # {f.professor: {f.departamento: {f.inicio: f.fim}}}
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Musculação' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                # {h.professor: {h.data: {h.departamento: h.horas}}}
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Musculação' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                # {a.professor: {a.data: a.departamento}}
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Musculação' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                # [datas de feriado formato dt]
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                # {e.professor: {e.data: {e.departamento: e.horas}}}
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Musculação' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))
        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Ginástica'
    novalinha += 1
    for i in ginastica:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Ginástica'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Ginástica' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                    for numb in subs:
                        for nome in subs[numb]:
                            for substituto in subs[numb][nome]:
                                for depart in subs[numb][nome][substituto]:
                                    for dia in subs[numb][nome][substituto][depart]:
                                        if depart == 'Ginástica' and nome == i:
                                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                    dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                                plan1.cell(column=cell.column, row=novalinha).fill = falta
                                        if depart == 'Ginástica' and substituto == i:
                                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                    dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                                plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                    column=cell.column, row=novalinha).value + float(
                                                    str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                                plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Ginástica' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                      month=int(str(plan1.cell(column=cell.column,
                                                                                               row=cell.row + 1).value).split(
                                                                              '/')[1]),
                                                                      year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Ginástica' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Ginástica' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Ginástica' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Ginástica' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))

        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Kids'
    novalinha += 1
    for i in kids:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Kids'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Kids' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                for numb in subs:
                    for nome in subs[numb]:
                        for substituto in subs[numb][nome]:
                            for depart in subs[numb][nome][substituto]:
                                for dia in subs[numb][nome][substituto][depart]:
                                    if depart == 'Kids' and nome == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).fill = falta
                                    if depart == 'Kids' and substituto == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                column=cell.column, row=novalinha).value + float(
                                                str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                            plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Kids' and nome == i and dt.strptime(dslg[nome][depart], '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dslg[nome][depart], '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                                     month=int(str(plan1.cell(
                                                                                             column=cell.column,
                                                                                             row=cell.row + 1).value).split(
                                                                                             '/')[1]),
                                                                                     year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Kids' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Kids' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Kids' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Kids' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))
        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Esportes'
    novalinha += 1
    for i in esportes:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Esportes'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Esportes' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                for numb in subs:
                    for nome in subs[numb]:
                        for substituto in subs[numb][nome]:
                            for depart in subs[numb][nome][substituto]:
                                for dia in subs[numb][nome][substituto][depart]:
                                    if depart == 'Esportes' and nome == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).fill = falta
                                    if depart == 'Esportes' and substituto == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                column=cell.column, row=novalinha).value + float(
                                                str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                            plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Esportes' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                      month=int(str(plan1.cell(column=cell.column,
                                                                                               row=cell.row + 1).value).split(
                                                                              '/')[1]),
                                                                      year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Esportes' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Esportes' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Esportes' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Esportes' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))

        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Cross Cia'
    novalinha += 1
    for i in cross:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Cross Cia'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Cross Cia' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                for numb in subs:
                    for nome in subs[numb]:
                        for substituto in subs[numb][nome]:
                            for depart in subs[numb][nome][substituto]:
                                for dia in subs[numb][nome][substituto][depart]:
                                    if depart == 'Cross Cia' and nome == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).fill = falta
                                    if depart == 'Cross Cia' and substituto == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                column=cell.column, row=novalinha).value + float(
                                                str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                            plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Cross Cia' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                      month=int(str(plan1.cell(column=cell.column,
                                                                                               row=cell.row + 1).value).split(
                                                                              '/')[1]),
                                                                      year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Cross Cia' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Cross Cia' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Cross Cia' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Cross Cia' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))

        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'PNT'
    novalinha += 1
    for i in pnt:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'PNT')/2)
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'PNT' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                for numb in subs:
                    for nome in subs[numb]:
                        for substituto in subs[numb][nome]:
                            for depart in subs[numb][nome][substituto]:
                                for dia in subs[numb][nome][substituto][depart]:
                                    if depart == 'PNT' and nome == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).fill = falta
                                    if depart == 'PNT' and substituto == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                column=cell.column, row=novalinha).value + float(
                                                str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                            plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'PNT' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                      month=int(str(plan1.cell(column=cell.column,
                                                                                               row=cell.row + 1).value).split(
                                                                              '/')[1]),
                                                                      year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'PNT' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'PNT' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'PNT' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'PNT' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))

        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1
    
    plan1[f'A{novalinha}'].value = 'Mensalistas'
    novalinha += 1
    for i in mensalistas:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_segunda(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_terca(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quarta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_quinta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sexta(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_sabado(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somar_aulas_de_domingo(str(plan1.cell(column=cell.column, row=4).value) + f'/{descobrir_ano(str(plan1.cell(column=cell.column, row=4).value))}', i, 'Mensalistas')/2)
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Mensalistas' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                for numb in subs:
                    for nome in subs[numb]:
                        for substituto in subs[numb][nome]:
                            for depart in subs[numb][nome][substituto]:
                                for dia in subs[numb][nome][substituto][depart]:
                                    if depart == 'Mensalistas' and nome == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).fill = falta
                                    if depart == 'Mensalistas' and substituto == i:
                                        if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                                dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                            plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(
                                                column=cell.column, row=novalinha).value + float(
                                                str(subs[numb][nome][substituto][depart][dia]).replace(',', '.'))
                                            plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Mensalistas' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(
                                        str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                      month=int(str(plan1.cell(column=cell.column,
                                                                                               row=cell.row + 1).value).split(
                                                                              '/')[1]),
                                                                      year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Mensalistas' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(
                                            str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]),
                                                                           month=int(str(plan1.cell(column=cell.column,
                                                                                                    row=cell.row + 1).value).split(
                                                                                   '/')[1]),
                                                                           year=dt.today().year) <= dt.strptime(
                                            fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Mensalistas' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(complem[nome][dia][depart]).replace(',', '.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Mensalistas' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                    dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Mensalistas' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(
                                        dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column,
                                                                                                     row=novalinha).value + float(
                                        str(escal[nome][dia][depart]).replace(',', '.'))
        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    for i, coluna in enumerate(plan1.columns):
        max_length = 0
        column = coluna[0].column_letter
        for cell in coluna:
            try:
                if len(str(cell.value)) > max_length and len(str(cell.value)) != 18:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 1
        plan1.column_dimensions[column].width = adjusted_width

    for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
        for cell in row:
            if plan1.cell(column=cell.column, row=cell.row).value == 'Total':
                plan1.column_dimensions[openpyxl.utils.cell.get_column_letter(cell.column)].width = 8

    # formatar coloração de fds
    for row in plan1.iter_cols(min_row=6, min_col=2, max_row=150, max_col=2):
        for cell in row:
            if cell.value is None and plan1.cell(column=cell.column, row=cell.row - 1).value is not None:
                    ultima_linha = cell.row

    for itens in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
        for cell in itens:
            if cell.value != 'Total':
                if cell.value == 'sáb' or cell.value == 'dom':
                    letras = openpyxl.utils.cell.get_column_letter(cell.column)
                    for numero in range(3, ultima_linha):
                        plan1[f'{letras}{numero}'].fill = fds

    plan1['C1'].fill = atestado
    plan1['D1'].value = 'Atestado'
    plan1['C2'].fill = falta
    plan1['D2'].value = 'Falta'
    plan1['F1'].fill = ferias
    plan1['G1'].value = 'Férias'
    plan1['F2'].fill = feriado
    plan1['G2'].value = 'Feriado'
    plan1['I1'].fill = deslig
    plan1['J1'].value = 'Desligamento'
    plan1['I2'].fill = subst
    plan1['J2'].value = 'Substituiu'
    plan1['M1'].fill = comple
    plan1['N1'].value = 'Horas Complementares'
    grade.save(pasta_pgto + f'\\Grade {fechamento.month}-{fechamento.year}.xlsx')


def somar_aulas_de_segunda(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulasseg = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Segunda') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasseg:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas


def somar_aulas_de_terca(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulaster = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Terça') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulaster:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas


def somar_aulas_de_quarta(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulasqua = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Quarta') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasqua:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas


def somar_aulas_de_quinta(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulasqui = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Quinta') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasqui:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas


def somar_aulas_de_sexta(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulassex = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Sexta') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulassex:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas


def somar_aulas_de_sabado(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulassab = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Sábado') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulassab:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas


def somar_aulas_de_domingo(diaplan: str, nome: str, depto: str) -> float:
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    aulasdom = session.query(Aulas).filter_by(professor=nome).filter_by(status='Ativa').filter_by(diadasemana='Domingo') \
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasdom:
        if aula.fimgrade is not None:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y') < dt.strptime(aula.fimgrade, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
        else:
            if dt.strptime(aula.iniciograde, '%d/%m/%Y') <= dt.strptime(diaplan, '%d/%m/%Y'):
                hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
                somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
                somadia = round(somadia, 2)
                somas += somadia
    session.close()
    return somas * 2


def salvar_planilha_soma_final(compet: int, year: int):
    mes = str(compet).zfill(2)
    ano = year
    mesext = {'01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR', '05': 'MAI', '06': 'JUN',
              '07': 'JUL', '08': 'AGO', '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'}
    pasta_pgto = rf'\\192.168.0.250\rh\01 - RH\01 - Administração.Controles\04 - Folha de Pgto\{ano}\{mes} - {mesext[mes]}\Grades e Comissões'
    try:
        os.makedirs(pasta_pgto)
    except FileExistsError:
        pass
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    folhadehoje = Folha(compet, list(listar_aulas_ativas(compet)), listar_departamentos_ativos())

    # criar dicionario somaaulas[professor][departamento][aulanome+valor][somadeaulas] que será usado na planilha de total
    somaaulas = {}
    for i in listar_professores_ativos():
        somaaulas[i] = {}
        for d in listar_departamentos_ativos():
            somaaulas[i][d] = {}
    for aulas in listar_aulas_ativas(compet):
        somaaulas[aulas.professor][aulas.departamento][aulas.nome + f' ({aulas.valor})'] = round(
            somar_horas_professor(folhadehoje, aulas.professor, aulas.departamento, aulas.nome, compet), 2)
        dictchav = list(somaaulas.keys())
        dictchav.sort()
        somafinal = {i: somaaulas[i] for i in dictchav}
    arquivo = os.path.relpath(rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\static\files\Somafinal.xlsx')
    plan = l_w(arquivo, read_only=False)
    folha = plan['Planilha1']
    folha['A1'].value = 'Matrícula'
    folha['B1'].value = 'Nome'
    folha['C1'].value = 'Aula e Valor'
    folha['D1'].value = 'Horas'
    x = 2
    for i in somafinal:
        matr = session.query(Aulas).filter_by(professor=str(i)).filter_by(status='Ativa').first()
        folha[f'A{x}'].value = int(matr.matrprof)
        folha[f'B{x}'].value = str(i)
        for sub in somafinal[i]:
            if somafinal[i][sub] == {}:
                pass
            else:
                for sub2 in somafinal[i][sub]:
                    folha[f'C{x}'].value = str(sub2)
                    folha[f'D{x}'].value = float(str(somafinal[i][sub][sub2]))
                    x += 1
    plan.save(pasta_pgto + f'\\Somafinal mes {compet}.xlsx')
    salvar_planilha_grade_horaria(somafinal, compet, ano)
    substitutos = {}
    complementares = {}
    feriasl = {}
    desligadosl = {}
    planilha = l_w(pasta_pgto + f'\\Grade {compet}-{ano}.xlsx')
    aba = planilha['Planilha1']
    for row in aba.iter_cols(min_row=3, min_col=3, max_row=150, max_col=35):
        for cell in row:
            if cell.fill == ferias:
                for i in range(1, 150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                feriasl[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}
            if cell.fill == deslig:
                for i in range(1, 150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                desligadosl[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}
            if cell.fill == subst:
                for i in range(1, 150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                substitutos[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}
            if cell.fill == comple:
                for i in range(1, 150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                complementares[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}

    planilha2 = l_w(pasta_pgto + f'\\Somafinal mes {compet}.xlsx', read_only=False)
    aba2 = planilha2['Planilha1']
    for pessoa in feriasl:
        for depart in feriasl[pessoa]:
            for cll in aba2['B']:
                if cll.value is not None:
                    if cll.value == pessoa:
                        aba2.cell(column=4, row=cll.row, value=float(feriasl[pessoa][depart]))
    for pessoa in desligadosl:
        for depart in desligadosl[pessoa]:
            for cell in aba2['B']:
                if cell.value is not None:
                    if cell.value == pessoa:
                        aba2.cell(column=4, row=cell.row).value = float(desligadosl[pessoa][depart])
    for pessoa in substitutos:
        for depart in substitutos[pessoa]:
            for cell in aba2['B']:
                if cell.value is not None:
                    if cell.value == pessoa:
                        aba2.cell(column=4, row=cell.row).value = float(substitutos[pessoa][depart])
    for pessoa in complementares:
        for depart in complementares[pessoa]:
            for cell in aba2['B']:
                if cell.value is not None:
                    if cell.value == pessoa:
                        aba2.cell(column=4, row=cell.row).value = float(complementares[pessoa][depart])
    for i, coluna in enumerate(aba2.columns):
        max_length = 0
        column = coluna[0].column_letter
        for cell in coluna:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 1
        aba2.column_dimensions[column].width = adjusted_width
    planilha2.save(pasta_pgto + f'\\Somafinal mes {compet}.xlsx')

    planilha3 = l_w(pasta_pgto + f'\\Grade {compet}-{ano}.xlsx', read_only=False)
    aba3 = planilha3['Planilha1']
    for row in aba3.iter_cols(min_row=6, min_col=2, max_row=150, max_col=2):
        for cell in row:
            if cell.value is None and aba3.cell(column=cell.column, row=cell.row - 1).value is not None:
                    ultima_linha = cell.row - 1

    for row in aba3.iter_cols(min_row=6, min_col=3, max_row=ultima_linha, max_col=35):
        for cell in row:
            if cell.value == 0:
                cell.value = ''

    for linha_dptos in aba3.iter_cols(min_row=4, min_col=1, max_row=ultima_linha, max_col=1):
        for celula in linha_dptos:
            if celula.value == 'Mensalistas':
                linha_mensalistas = int(celula.row)

    for linha_tt in aba3.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
        for celula in linha_tt:
            if celula.value == 'Total':
                coluna_total = int(celula.column)
                letratt = openpyxl.utils.cell.get_column_letter(celula.column - 1)

    for row in aba3.iter_cols(min_row=linha_mensalistas + 1, min_col=3, max_row=ultima_linha, max_col=coluna_total):
        for c in row:
            if c.value != '':
                if aba3.cell(column=c.column, row=3).value == 'Total':
                    aba3.cell(column=c.column, row=c.row, value=f'=COUNTA(C{c.row}:{letratt}{c.row})')
                else:
                    c.value = 'P'
            if c.fill == ferias:
                c.value = 'F'
            if c.fill == atestado:
                c.value = 'A'
            if c.fill == deslig:
                c.value = ''
            if c.fill == subst:
                c.value = 'D'
            if c.fill == comple:
                c.value = 'D'
            if c.fill == falta:
                c.value = 'F'
    top_rows = aba3['AJ5']
    aba3.freeze_panes = top_rows
    planilha3.save(pasta_pgto + f'\\Grade {compet}-{ano}.xlsx')
    tkinter.messagebox.showinfo(
        title='Grade ok!',
        message=f'Grade do mês {compet} salva com sucesso!'
    )


def lancar_ferias(nome, depto, inicio, fim):
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    pessoa = sessioncol.query(Colaborador).filter_by(nome=nome).order_by(Colaborador.matricula.desc()).first()
    matricula = pessoa.matricula

    fer = Ferias(professor=nome, matrprof=matricula, departamento=depto, inicio=inicio, fim=fim)
    session.add(fer)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Férias ok!', 'Férias lançadas com sucesso!')


def lancar_atestado(nome, depto, data):
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()

    professor = sessioncol.query(Colaborador).filter_by(nome=nome).order_by(Colaborador.matricula.desc()).first()
    matricula = professor.matricula

    atest = Atestado(professor=nome, matrprof=matricula, departamento=depto, data=data)
    session.add(atest)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Atestado salvo!', 'Atestado salvo com sucesso!')


def lancar_desligamento(nome, depto, data):
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()

    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()

    professor = sessioncol.query(Colaborador).filter_by(nome=nome).order_by(Colaborador.matricula.desc()).first()
    matricula = professor.matricula

    deslig = Desligados(professor=nome, matrprof=matricula, departamento=depto, datadesligamento=data)
    session.add(deslig)
    session.commit()
    session.close()

    # se desligamento antes do iniício da folha: tornar inativas as aulas da pessoa
    tkinter.messagebox.showinfo('Desligamento ok!', 'Desligamento lançado com sucesso!')


def lancar_substit(substituido, substituto, departamento, aula, data, horas):
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    profsubstituido = sessioncol.query(Colaborador).filter_by(nome=substituido).order_by(Colaborador.matricula.desc()).first()
    matrsubsido = profsubstituido.matricula
    profsubstituto = sessioncol.query(Colaborador).filter_by(nome=substituto).order_by(Colaborador.matricula.desc()).first()
    matrsubstuto = profsubstituto.matricula
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    subs = Substituicao(professorsubst=substituido, matrprof=matrsubsido, substituto=substituto,
                        matrsubstituto=matrsubstuto, departamento=departamento, aula=aula, data=data,
                        horas=horas)
    session.add(subs)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Substituição ok!', 'Substituição lançada com sucesso!')


def lancar_hrscomple(nome, departamento, aula, data, horas):
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    pessoa = sessioncol.query(Colaborador).filter_by(nome=nome).order_by(Colaborador.matricula.desc()).first()
    matricula = pessoa.matricula
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    hrcomp = Hrcomplement(professor=nome, matrprof=matricula, departamento=departamento, horas=horas, aula=aula, data=data)
    session.add(hrcomp)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Horas Salvas!', 'Horas complementares salvas com sucesso!')


def lancar_faltas(nome, depto, data, hrs):
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    pessoa = sessioncol.query(Colaborador).filter_by(nome=nome).order_by(Colaborador.matricula.desc()).first()
    matricula = pessoa.matricula

    falt = Faltas(professor=nome, matrprof=matricula, departamento=depto, data=data, horas=hrs)
    session.add(falt)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Falta Salva!', 'Falta salva com sucesso!')


def lancar_novaaula(nomeprof, depto, nomeaula, diasemana, inicio, fim, valor):
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    inicio = inicio + ':00'
    fim = fim + ':00'
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    pessoa = sessioncol.query(Colaborador).filter_by(nome=nomeprof).order_by(Colaborador.matricula.desc()).first()
    matricula = pessoa.matricula
    hj = dt.today()
    aula = Aulas(nome=nomeaula, professor=nomeprof, departamento=depto, diadasemana=diasemana, inicio=inicio,
                 fim=fim, valor=valor, status='Ativa', iniciograde=dt.strftime(hj, '%d/%m/%Y'), matrprof=matricula)
    session.add(aula)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Horário Salvo!', 'Novo horário lançado com sucesso!')


def lancar_escala(nome, departamento, aula, data, horas):
    sessionscol = sessionmaker(bind=engine)
    sessioncol = sessionscol()
    pessoa = sessioncol.query(Colaborador).filter_by(nome=nome).order_by(Colaborador.matricula.desc()).first()
    matricula = pessoa.matricula
    sessions = sessionmaker(bind=enginefolha)
    session = sessions()
    esc = Escala(professor=nome, matrprof=matricula, departamento=departamento, horas=horas, aula=aula, data=data)
    session.add(esc)
    session.commit()
    session.close()
    tkinter.messagebox.showinfo('Escala Salva!', 'Escala salva com sucesso!')


def salvar_banco_aulas():
    sessions = sessionmaker(enginefolha)
    session = sessions()
    aula = session.query(Aulas).filter_by(status='Ativa').order_by(Aulas.professor).all()
    wb = l_w(rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\static\files\Plan.xlsx', read_only=False)
    sh = wb['Planilha1']
    x = 2
    for a in aula:
        sh[f'A{x}'].value = a.numero
        sh[f'B{x}'].value = a.professor
        sh[f'C{x}'].value = a.nome
        sh[f'D{x}'].value = a.departamento
        sh[f'E{x}'].value = a.diadasemana
        sh[f'F{x}'].value = a.inicio
        sh[f'G{x}'].value = a.fim
        sh[f'H{x}'].value = a.valor
        sh[f'I{x}'].value = a.iniciograde
        sh[f'J{x}'].value = a.matrprof
        x += 1
    sh[f'A1'].value = 'Número'
    sh[f'B1'].value = 'Professor'
    sh[f'C1'].value = 'Nome da Aula'
    sh[f'D1'].value = 'Departamento'
    sh[f'E1'].value = 'Dia da semana'
    sh[f'F1'].value = 'Início'
    sh[f'G1'].value = 'Fim'
    sh[f'H1'].value = 'Valor'
    sh[f'I1'].value = 'Início Grade'
    sh[f'J1'].value = 'Matrícula'
    wb.save(rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\static\files\CadastroAulas.xlsx')
    tkinter.messagebox.showinfo('Banco de aulas atualizado!', 'Aulas ativas salvas em excel com sucesso!')


def salvar_plan_lancamentos(comp):
    hj = dt.today()
    mes = str(comp).zfill(2)
    ano = hj.year
    mesext = {'01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR', '05': 'MAI', '06': 'JUN',
              '07': 'JUL', '08': 'AGO', '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'}
    pasta_pgto = rf'\\192.168.0.250\rh\01 - RH\01 - Administração.Controles\04 - Folha de Pgto\{ano}\{mes} - {mesext[mes]}\Grades e Comissões'

    wb = l_w(rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\static\files\Plan.xlsx', read_only=False)
    del wb['Planilha1']
    wb.create_sheet('Fluxo')
    wb.create_sheet('DeletarFerias')
    wb.create_sheet('Faltas')
    wb.create_sheet('Compl Est')
    wb.create_sheet('Horistas')
    wb.create_sheet('Comissoes')
    wb.create_sheet('Plano')
    wb.create_sheet('Adiantamento')
    wb.create_sheet('DescontoVT')
    wb.save(pasta_pgto + rf'\Lancamentos - {comp}.xlsx')

    tkinter.messagebox.showinfo('Planilha ok!', f'Planilha "Lançamentos" mês {comp} salva com sucesso.')


def inativar_aulas(aulas: list):
    sessions = sessionmaker(enginefolha)
    session = sessions()
    for aula in aulas:
        a = session.query(Aulas).filter_by(numero=aula).first()
        a.status = 'Inativa'
        session.commit()
    tkinter.messagebox.showinfo(title='Aulas inativas!',message='Aulas inativas com sucesso!')


def backup_bancos(bancoapp: str, bancoautomacao: str, sentido: int):
    """
    Backup databases from different aplications.
    :param bancoapp: db appCia
    :param bancoautomacao: db AutomaçãoCia
    :return:
    """
    # falta copiar todos os demais dados para ambos DB (faltas, atestados, subst, deslig, hrcompl etc)
    if bancoapp == '' or bancoautomacao == '':
        print(bancoapp)
        print(bancoautomacao)
        tkinter.messagebox.showinfo(title='Escolha os dois arquivos!', message='Um ou os dois arquivos não foram definidos!')
    else:
        # backup DB automaçao e DB app em pasta com hash e data
        text = str(datetime.datetime.now())
        hash_object = hashlib.sha1(text.encode())
        shutil.copyfile(bancoapp, rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\bkp\aulas_appcia_bkp {dt.strftime(dt.today(),"%d.%m.%Y")} - {hash_object.hexdigest()}.db')
        shutil.copyfile(bancoautomacao, rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\bkp\aulas_autom_bkp {dt.strftime(dt.today(), "%d.%m.%Y")} - {hash_object.hexdigest()}.db')

        # conectar no DB appCia e passar informações para DB automação
        arqf = os.path.abspath(bancoapp)
        enginef = create_engine('sqlite+pysqlite:///' + arqf, echo=True, future=True)
        metadata_obj = MetaData()
        Base = declarative_base()

        class Aulas(Base):
            __tablename__ = "aulas"
            numero = Column(Integer, primary_key=True)
            nome = Column(String, nullable=False)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            diadasemana = Column(String, nullable=False)
            inicio = Column(String, nullable=False)
            fim = Column(String, nullable=False)
            valor = Column(String, nullable=False)
            status = Column(String, nullable=False)
            iniciograde = Column(String, nullable=False)
            fimgrade = Column(String, nullable=True)

        class Faltas(Base):
            __tablename__ = "faltas"
            numero = Column(Integer, primary_key=True)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            data = Column(String, nullable=False)
            horas = Column(String, nullable=False)

        class Ferias(Base):
            __tablename__ = 'ferias'
            numero = Column(Integer, primary_key=True)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            inicio = Column(String, nullable=False)
            fim = Column(String, nullable=False)

        class Atestado(Base):
            __tablename__ = 'atestado'
            numero = Column(Integer, primary_key=True)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            data = Column(String, nullable=False)

        class Substituicao(Base):
            __tablename__ = 'substituicao'
            numero = Column(Integer, primary_key=True)
            professorsubst = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            substituto = Column(String, nullable=False)
            matrsubstituto = Column(String, nullable=False)
            departamento = Column(String, nullable=False)
            aula = Column(String, nullable=False)
            data = Column(String, nullable=False)
            horas = Column(String, nullable=False)

        class Desligados(Base):
            __tablename__ = 'desligados'
            numero = Column(Integer, primary_key=True)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            datadesligamento = Column(String, nullable=False)

        class Escala(Base):
            __tablename__ = 'escala'
            numero = Column(Integer, primary_key=True)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            aula = Column(String, nullable=False)
            horas = Column(String, nullable=False)
            data = Column(String, nullable=False)

        class Hrcomplement(Base):
            __tablename__ = 'hrcomplementar'
            numero = Column(Integer, primary_key=True)
            professor = Column(String, nullable=False)
            matrprof = Column(Integer, nullable=False)
            departamento = Column(String, nullable=False)
            aula = Column(String, nullable=False)
            horas = Column(String, nullable=False)
            data = Column(String, nullable=False)

        Base.metadata.create_all(enginef)
        if sentido == 1:
            # AppCia -> Automação
            sessionsapp = sessionmaker(enginef)
            sessionapp = sessionsapp()
            
            sessions = sessionmaker(enginefolha)
            session = sessions()
            
            aulasapp = sessionapp.query(Aulas).all()
            for aulapp in aulasapp:
                aulaaut = session.query(Aulas).filter_by(numero=aulapp.numero).first()
                if aulaaut:
                    aulaaut.nome = aulapp.nome
                    aulaaut.professor = aulapp.professor
                    aulaaut.matrprof = aulapp.matrprof
                    aulaaut.departamento = aulapp.departamento
                    aulaaut.diadasemana = aulapp.diadasemana
                    aulaaut.inicio = aulapp.inicio
                    aulaaut.fim = aulapp.fim
                    aulaaut.valor = aulapp.valor
                    aulaaut.status = aulapp.status
                    aulaaut.iniciograde = aulapp.iniciograde
                    aulaaut.fimgrade = aulapp.fimgrade
                    session.commit()
                else:
                    aula = Aulas(numero=aulapp.numero, nome=aulapp.nome, professor=aulapp.professor,
                                 matrprof=aulapp.matrprof, departamento=aulapp.departamento,
                                 diadasemana=aulapp.diadasemana, inicio=aulapp.inicio, fim=aulapp.fim,
                                 valor=aulapp.valor, status=aulapp.status, iniciograde=aulapp.iniciograde,
                                 fimgrade=aulapp.fimgrade)
                    session.add(aula)
                    session.commit()
            session.close()
            sessionapp.close()
            tkinter.messagebox.showinfo(title='Download ok!', message='Aulas copiadas do AppCia para AutomaçãoCia!')
        else:
            # Automação -> AppCia
            sessionsapp = sessionmaker(enginef)
            sessionapp = sessionsapp()

            sessions = sessionmaker(enginefolha)
            session = sessions()

            aulasaut = session.query(Aulas).all()
            for aulaut in aulasaut:
                aulaapp = sessionapp.query(Aulas).filter_by(numero=aulaut.numero).first()
                if aulaapp:
                    aulaapp.nome = aulaut.nome
                    aulaapp.professor = aulaut.professor
                    aulaapp.matrprof = aulaut.matrprof
                    aulaapp.departamento = aulaut.departamento
                    aulaapp.diadasemana = aulaut.diadasemana
                    aulaapp.inicio = aulaut.inicio
                    aulaapp.fim = aulaut.fim
                    aulaapp.valor = aulaut.valor
                    aulaapp.status = aulaut.status
                    aulaapp.iniciograde = aulaut.iniciograde
                    aulaapp.fimgrade = aulaut.fimgrade
                    sessionapp.commit()
                else:
                    aula = Aulas(numero=aulaut.numero, nome=aulaut.nome, professor=aulaut.professor,
                                 matrprof=aulaut.matrprof, departamento=aulaut.departamento,
                                 diadasemana=aulaut.diadasemana, inicio=aulaut.inicio, fim=aulaut.fim,
                                 valor=aulaut.valor, status=aulaut.status, iniciograde=aulaut.iniciograde,
                                 fimgrade=aulaut.fimgrade)
                    sessionapp.add(aula)
                    sessionapp.commit()
            session.close()
            sessionapp.close()
            tkinter.messagebox.showinfo(title='Upload ok!', message='Aulas copiadas do AutomaçãoCia para AppCia!')


def previa_folha():
    """
    Email workers with a preview of payment calculations.
    :return:
    """
    mesext = {'01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR', '05': 'MAI', '06': 'JUN',
              '07': 'JUL', '08': 'AGO', '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'}

    # deletar todos arquivos zip destinatarios
    for arquivo_destinatario in os.listdir(pasta_dexion):
        apagar = os.path.join(pasta_dexion, arquivo_destinatario)
        if os.path.isfile(apagar):
            if 'destinatario' in arquivo_destinatario:
                os.remove(os.path.join(pasta_dexion, arquivo_destinatario))

    # set up smtp connection
    email_remetente = em_rem
    senha = k1
    smtp = smtplib.SMTP(host=host, port=port)
    smtp.starttls()
    smtp.login(email_remetente, senha)

    # connect db
    sessions = sessionmaker(engine)
    session = sessions()

    # Excel
    excel = client.Dispatch('Excel.Application')
    excel.visible = 0

    # deletar pasta Protocolos
    try:
        shutil.rmtree(pasta_dexion + r'\Protocolos')
    except FileNotFoundError:
        pass

    # descompactar arquivo do contracheque
    caminho = pasta_dexion
    for filename in os.listdir(caminho):
        cam_arq = os.path.join(caminho, filename)
        file = filename.replace('.zip', '').replace('.pdf', '')
        filepdf = filename.replace('.zip', '.pdf')
        if os.path.isfile(cam_arq) and filename.endswith('.zip'):
            cam, competencia, data_pgto, matricula = cam_arq.split(',')
            data_pgto = data_pgto.strip()
            mes, ano = competencia.split('-')
            mes = mes.strip()
            ano = ano.strip()
            matricula = matricula.replace(').zip', '')
            matricula = int(matricula)
            with zipfile.ZipFile(cam_arq, 'r') as zip_ref:
                zip_ref.extract(filepdf, caminho)

        # converter pdf em imagem
        images = convert_from_path(os.path.join(caminho, filepdf))
        for i in range(len(images)):
            images[i].save(os.path.join(caminho, file) + '.jpg', 'JPEG')
        os.remove(os.path.join(caminho, filepdf))

        # cortar a imagem e salvar img ajustada
        rect_size = (875, 88)
        rect_pos = (9, 660)
        im = Image.open(os.path.join(caminho, file + '.jpg'))
        area = (72, 304, 1370, 1063)
        img = im.crop(area)
        im.close()
        img.save(os.path.join(caminho, file + ' cortada' + '.jpg'))

        # escrever na imagem cortada
        img = Image.open(os.path.join(caminho, file + ' cortada' + '.jpg')).convert("RGBA")
        rect = Image.new("RGBA", rect_size, (255, 255, 255, 255))
        img.paste(rect, rect_pos)
        f = ImageFont.truetype("arial.ttf", 150)
        font2 = ImageFont.truetype("arial.ttf", 30)
        txt = Image.new('L', (1000, 1400))
        d = ImageDraw.Draw(txt)
        d.text((0, -20), f'      PRÉVIA\n         DE\n PAGAMENTO', font=f, fill=125)
        w = txt.rotate(20, expand=True)
        drawing = ImageDraw.Draw(img)
        drawing.text((9, 570), 'Esse documento não é o contracheque oficial.\nRelatório para simples conferência.', (255, 0, 0),
                     font=font2)
        img.paste(w, (120, -20), w)
        imagem = img.convert('RGB')
        img.close()
        imagem.save(os.path.join(caminho, file + '.jpg'))
        os.remove(os.path.join(caminho, file + ' cortada' + '.jpg'))

        # # printar planilha de grade com as hrs
        competencia = str(mes).zfill(2) + '-' + str(dt.today().year)
        pagamento = dt.strftime(dt.strptime(data_pgto, '%d-%m-%Y'), '%d-%m-%Y')
        planfolha = rf'\\192.168.0.250\rh\01 - RH\01 - Administração.Controles\04 - Folha de Pgto\{ano}\{mes} - {mesext[mes]}\Grades e Comissões\Grade {mes}-{ano}.xlsx'
        pessoa = session.query(Colaborador).filter_by(matricula=matricula).order_by(Colaborador.matricula.desc()).first()
        plan = excel.Workbooks.Open(planfolha)
        folha = plan.Sheets['Planilha1']
        # salvar linhas e colunas bases das datas de folha
        copyrange = folha.Range('C1:AI4')
        copyrange.CopyPicture(Format=2)
        ImageGrab.grabclipboard().save(pasta_dexion + rf'\Grade (A - Mensal, {mes}-{ano}, {data_pgto}, {str(matricula).zfill(6)}).jpg')
        # descobrir a linha (ou linhas) que está o nome e incluir em lista
        linhas = []
        plan = l_w(planfolha)
        sh = plan['Planilha1']
        for row in sh.iter_rows(min_row=1, min_col=2, max_row=150, max_col=2):
            for cell in row:
                if cell.value == pessoa.nome:
                    linhas.append(sh.cell(row=cell.row, column=cell.column).row)
        size = 81
        for linha in linhas:
            copyrange = folha.Range(f'C{linha}:AI{linha}')
            copyrange.CopyPicture(Format=2)
            ImageGrab.grabclipboard().save(pasta_dexion + rf"\Grade {pessoa.nome} {linha}.jpg")
            image1 = Image.open(pasta_dexion + rf'\Grade (A - Mensal, {mes}-{ano}, {data_pgto}, {str(matricula).zfill(6)}).jpg')
            image2 = Image.open(pasta_dexion + rf"\Grade {pessoa.nome} {linha}.jpg")
            new_image = Image.new('RGB', (1772, 150), (250, 250, 250))
            new_image.paste(image1, (0, 0))
            new_image.paste(image2, (0, size))
            new_image.save(pasta_dexion + rf'\Grade (A - Mensal, {mes}-{ano}, {data_pgto}, {str(matricula).zfill(6)}).jpg')
            os.remove(pasta_dexion + rf"\Grade {pessoa.nome} {linha}.jpg")
            size += 21
        excel.Quit()
        # apagar arquivo zip:
        os.remove(os.path.join(caminho, filename))

        # # fazer pdf com os dois prints
        pdf = FPDF()
        imagelist = [os.path.join(caminho, file + '.jpg'),
                     pasta_dexion + rf'\Grade (A - Mensal, {mes}-{ano}, {data_pgto}, {str(matricula).zfill(6)}).jpg']
        pdf.add_page()
        pdf.image(imagelist[0], 0, 0, 200, 100)
        pdf.image(imagelist[1], 5, 110, 255, 20)
        pdf.output(os.path.join(caminho, f'Prévia Folha - {pessoa.nome.title()}.pdf'))

        # deletar imagens
        os.remove(os.path.join(caminho, file + '.jpg'))
        os.remove(pasta_dexion + rf'\Grade (A - Mensal, {mes}-{ano}, {data_pgto}, {str(matricula).zfill(6)}).jpg')

        # enviar arquivo pdf por e-mail
        msg = MIMEMultipart('alternative')
        msg['From'] = email_remetente
        msg['To'] = pessoa.email
        msg['Subject'] = 'Prévia de Pagamento'
        arquivo = os.path.join(caminho, f'Prévia Folha - {pessoa.nome.title()}.pdf')
        if pessoa.tipo_contr == 'Horista':
            text = MIMEText(f'''Olá, {str(pessoa.nome).title().split(" ")[0]}!<br><br>
            Segue anexo pdf com a prévia do seu pagamento do próximo mês.<br>
            Confira todas as rubricas lançadas, o total de horas, os valores e a grade abaixo da prévia.<br>
            As alterações na prévia podem ser feitas até o dia 25 desse mês.<br>
            Qualquer dúvida, estou à disposição.<br><br>
            Atenciosamente,<br>
            <img src="cid:image1">''', 'html')
        else:
            text = MIMEText(f'''Olá, {str(pessoa.nome).title().split(" ")[0]}!<br><br>
            Segue anexo pdf com a prévia do seu pagamento da folha desse mês.<br>
            Confira o total de dias, as faltas, os valores lançados e a grade abaixo da prévia.<br>
            As alterações na prévia podem ser feitas até o dia 25 desse mês.<br>
            Qualquer dúvida, estou à disposição.<br><br>
            Atenciosamente,<br>
            <img src="cid:image1">''', 'html')
        # set up the parameters of the message
        msg.attach(text)
        image = MIMEImage(
            open(rf'C:\Users\{os.getlogin()}\PycharmProjects\AutomacaoCia\src\models\static\imgs\assinatura.png',
                 'rb').read())
        image.add_header('Content-ID', '<image1>')
        msg.attach(image)
        # attach pdf file
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(arquivo, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename='Prévia de pagamento.pdf')
        msg.attach(part)
        smtp.sendmail(email_remetente, pessoa.email, msg.as_string())
        del msg

        # deletar todos arquivos criados e arquivo zip e da pasta do dexion
        os.remove(os.path.join(caminho, f'Prévia Folha - {pessoa.nome.title()}.pdf'))
    smtp.quit()
    tkinter.messagebox.showinfo(title='Prévias enviadas!', message='Prévias da folha enviadas com sucesso!')
