from datetime import datetime as dt
import locale
from openpyxl import load_workbook as l_w
import os
import pyautogui as pa
from PIL import ImageGrab, Image
from src.models.models import Colaborador, engine
from sqlalchemy.orm import sessionmaker
from src.models.dados_servd import pasta_dexion, pasta_estag, pasta_func
import tkinter.filedialog
from tkinter import messagebox
import tkinter.filedialog
import time as t
import win32com.client as client
import zipfile
from zipfile import ZipFile

locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')


def salvar_holerites():
    caminho = pasta_dexion
    for filename in os.listdir(caminho):
        f = os.path.join(caminho, filename)
        if os.path.isfile(f) and filename.endswith('.zip'):
            cam, competencia, data_pgto, matricula = f.split(',')
            mes, ano = competencia.split('-')
            matricula = matricula.replace(').zip', '')
            matricula = int(matricula)
            sessions = sessionmaker(bind=engine)
            session = sessions()
            pessoa = session.query(Colaborador).filter_by(matricula=matricula).first()
            if pessoa:
                if pessoa.cargo == 'ESTAGIARIO' or pessoa.cargo == 'ESTAGIARIA':
                    pasta_funcional = pasta_estag + f'\\{str(pessoa.nome).strip()}'
                else:
                    pasta_funcional = pasta_func + f'\\{str(pessoa.nome).strip()}'
                try:
                    os.makedirs(pasta_funcional + f'\\Holerites\\{ano}\\{mes}')
                except FileExistsError:
                    pass
                try:
                    ZipFile(f, 'r').extractall(path=pasta_funcional + f'\\Holerites\\{ano}\\{mes}')
                except FileNotFoundError:
                    pass
    tkinter.messagebox.showinfo('Contracheques Salvos!', 'Contracheques salvos com sucesso!')


def incluir_grade_email_holerite(planfolha: str, comp: str, pgto: str):
    dia, mes, ano = comp.split('/')
    competencia = str(mes).zfill(2) + '-' + str(ano)
    pagamento = dt.strftime(dt.strptime(pgto, '%d/%m/%Y'), '%d-%m-%Y')
    pst_dexion = r'\\192.168.0.234\Dexion\Logistica\000001\Folha'

    nomes = []
    pl = l_w(planfolha)
    s = pl['Planilha1']
    for row in s.iter_rows(min_row=5, min_col=2, max_row=126, max_col=2):
        for cell in row:
            nomes.append(str(cell.value).strip())
            nomes = list(sorted(set(filter(None, nomes))))
    excel = client.Dispatch('Excel.Application')
    plan = excel.Workbooks.Open(planfolha)
    folha = plan.Sheets['Planilha1']
    excel.visible = 0
    sessions = sessionmaker(engine)
    session = sessions()
    for nome in nomes:
        pessoa = session.query(Colaborador).order_by(Colaborador.matricula.desc()).filter_by(nome=nome).first()
        if pessoa:
            matricula = pessoa.matricula
        else:
            matricula = ''
        # salvar linhas e colunas bases das datas de folha
        copyrange = folha.Range('A1:AI4')
        copyrange.CopyPicture(Format=2)
        ImageGrab.grabclipboard().save(pst_dexion + r"\Grade.png")
        # descobrir a linha (ou linhas) que está o nome e incluir em lista
        linhas = []
        plan = l_w(planfolha)
        sh = plan['Planilha1']
        for row in sh.iter_rows(min_row=1, min_col=2, max_row=126, max_col=2):
            for cell in row:
                if cell.value == nome:
                    linhas.append(sh.cell(row=cell.row, column=cell.column).row)
        size = 81
        for linha in linhas:
            copyrange = folha.Range(f'A{linha}:AI{linha}')
            copyrange.CopyPicture(Format=2)
            ImageGrab.grabclipboard().save(pst_dexion + rf"\Grade {nome} {linha}.png")
            image1 = Image.open(pst_dexion + r"\Grade.png")
            image2 = Image.open(pst_dexion + rf"\Grade {nome} {linha}.png")
            new_image = Image.new('RGB', (1772, 150), (250, 250, 250))
            new_image.paste(image1, (0, 0))
            new_image.paste(image2, (0, size))
            new_image.save(pst_dexion + r"\Grade.png")
            os.remove(pst_dexion + rf"\Grade {nome} {linha}.png")
            size += 21
        zipf = zipfile.ZipFile(
          pst_dexion + rf'\Recibo de Pagamento (A - Mensal, {competencia}, {pagamento}, {str(matricula).zfill(6)}).zip',
          'a'
        )
        try:
            zipf.write(pst_dexion + r"\Grade.png", os.path.basename(pst_dexion + r"\Grade.png"))
        except FileNotFoundError:
            pass
        zipf.close()
        os.remove(pst_dexion + r"\Grade.png")
    excel.Quit()
    tkinter.messagebox.showinfo('Sucesso!', 'Prints de grades adicionados aos arquivos ".zip" com sucesso!')


def emitir_contracheque():
    # code to automate the process of creation of documents
    wb = l_w('Contracheque.xlsx')
    sh = wb['Planilha1']
    x = 2
    while x <= len(sh['A']):
        competencia = str(sh[f'A{x}'].value).replace('/', '')
        pagamento = str(sh[f'B{x}'].value).replace('/', '')
        de = str(sh[f'C{x}'].value)
        ate = str(sh[f'D{x}'].value)
        caminho = \
            f'C:\\Users\\RH\\PycharmProjects\\AutomacaoCia\\Emissao de contracheques\\Contracheque {competencia}.pdf'
        pa.click(-816, 515), t.sleep(0.5), pa.write(competencia), pa.press('tab'), pa.write(pagamento), pa.press('tab')
        pa.write(de), pa.press('tab'), pa.write(ate), pa.click(-787, 731), t.sleep(4)
        pa.hotkey('ctrl', 's'), pa.write(caminho), t.sleep(0.5), pa.press('enter'), t.sleep(0.5), pa.click(-33, 132)
        x += 1
