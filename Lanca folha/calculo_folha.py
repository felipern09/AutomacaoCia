from datetime import datetime as dt, timedelta as td
from dateutil.relativedelta import relativedelta
from modelsfolha import Aulas, Faltas, Ferias, Hrcomplement, Atestado, Desligados, Escala, Substituicao, engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook as l_w
from openpyxl.styles import PatternFill, Font
import pandas as pd
import openpyxl.utils.cell
import holidays
import locale
locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
Sessions = sessionmaker(bind=engine)
session = Sessions()
locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')
data = int(input('Digite o mes da competência: '))


class Folha:
    def __init__(self, competencia, aulas, deptos):
        self.fechamento = dt(day=20, month=competencia, year=dt.today().year)
        self.compet = self.fechamento.month
        self.aulas = aulas
        self.dept = deptos


class Aula:
    def __init__(self, nome, prof, depart, diasem,inicio, fim, valorhr, iniciograde, fimgrade=''):
        self.nome = nome
        self.professor = prof
        self.departamento = depart
        self.dia = diasem
        self.inicio = dt.strptime(inicio, '%H:%M:%S')
        self.fim = dt.strptime(fim, '%H:%M:%S')
        self.valor = valorhr
        self.iniciograde = iniciograde
        self.fimgrade = fimgrade
        self.dsr = 1.1666


def somaaula(diasem, inic, fim, competencia, iniciograd, fimgrad):
    horas = fim - inic
    hr, minut, seg = str(horas).split(':')
    igrad = dt.strptime(iniciograd, '%d/%m/%Y')
    if fimgrad is not None:
        fgrad = dt.strptime(fimgrad, '%d/%m/%Y')
    else:
        fgrad = ''
    somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
    somadia = round(somadia, 2)
    soma = 0
    comp = dt(day=1, month=competencia, year=dt.today().year)
    inicio = dt(day=21, month=(comp - relativedelta(months=1)).month, year=(comp - relativedelta(months=1)).year)
    fechamento = dt(day=20, month=comp.month, year=comp.year)

    def intervalo(inicio, fechamento):
        for n in range(int((fechamento - inicio).days) + 1):
            yield inicio + td(n)

    if fgrad != '':
        if diasem == 'Segunda':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 0 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Terça':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 1 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Quarta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 2 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Quinta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 3 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Sexta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 4 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Sábado':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 5 and igrad <= dia <= fgrad:
                    soma += somadia
        if diasem == 'Domingo':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 6 and igrad <= dia <= fgrad:
                    soma += somadia
    else:
        if diasem == 'Segunda':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 0 and dia >= igrad:
                    soma += somadia
        if diasem == 'Terça':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 1 and dia >= igrad:
                    soma += somadia
        if diasem == 'Quarta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 2 and dia >= igrad:
                    soma += somadia
        if diasem == 'Quinta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 3 and dia >= igrad:
                    soma += somadia
        if diasem == 'Sexta':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 4 and dia >= igrad:
                    soma += somadia
        if diasem == 'Sábado':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 5 and dia >= igrad:
                    soma += somadia
        if diasem == 'Domingo':
            for dia in intervalo(inicio, fechamento):
                if dia.weekday() == 6 and dia >= igrad:
                    soma += somadia
    return round(soma, 2)


def aulasativas():
    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    aula = []
    for i, a in enumerate(aulasativasdb):
        aula.append(i)
        aula[i] = Aula(a.nome, a.professor, a.departamento, a.diadasemana, a.inicio, a.fim, a.valor, a.iniciograde, a.fimgrade)
        yield aula[i]


def deptosativos():
    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    departamentos = []
    for i, a in enumerate(aulasativasdb):
        departamentos.append(a.departamento)
        departamentos = list(set(departamentos))
    return departamentos


def profsativos():
    aulasativasdb = session.query(Aulas).filter_by(status='Ativa').all()
    professores = []
    for i, a in enumerate(aulasativasdb):
        professores.append(a.professor)
        professores = list(set(professores))
    return professores


def totaldafolha(folha):
    somatorio = 0
    for al in list(aulasativas()):
        somatorio += somaaula(al.dia, al.inicio, al.fim, data, al.iniciograde, al.fimgrade) * float(str(al.valor).replace(',', '.')) * al.dsr
    return round(somatorio, 2)


def somaprof(folha, prof, depto, nome):
    somahoras = 0
    for aula in folha.aulas:
        if aula.professor == prof and aula.departamento == depto and aula.nome == nome:
            somahoras += somaaula(aula.dia, aula.inicio, aula.fim, data, aula.iniciograde, aula.fimgrade)
    return somahoras


def faltas(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    falt = session.query(Faltas).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for f in falt:
        if inicio <= dt.strptime(f.data, '%d/%m/%Y') <= fim:
            if f.professor in dic:
                d2 = {f.professor: {f.data: {f.departamento: f.horas}}}
                dic[f.professor] = {**dic[f.professor], **d2[f.professor]}
            else:
                d2 = {f.professor: {f.data: {f.departamento: f.horas}}}
                dic = {**dic, **d2}
    return dic


def feriasf(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    fer = session.query(Ferias).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for f in fer:
        if inicio <= dt.strptime(f.inicio, '%d/%m/%Y') <= fim:
            if f.professor in dic:
                d2 = {f.professor: {f.departamento: {f.inicio: f.fim}}}
                dic[f.professor] = {**dic[f.professor], **d2[f.professor]}
            else:
                d2 = {f.professor: {f.departamento: {f.inicio: f.fim}}}
                dic = {**dic, **d2}
    return dic


def atestadof(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    ates = session.query(Atestado).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for a in ates:
        if inicio <= dt.strptime(a.data, '%d/%m/%Y') <= fim:
            if a.professor in dic:
                d2 = {a.professor: {a.data: a.departamento}}
                dic[a.professor] = {**dic[a.professor], **d2[a.professor]}
            else:
                d2 = {a.professor: {a.data: a.departamento}}
                dic = {**dic, **d2}
    return dic


def feriadof(comp):
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    # Get the Bank Holidays for the given country
    feriados = holidays.country_holidays('BR')
    # Create a list of dates between the start and end date
    intervalo_datas = pd.date_range(inicio, fim)
    # Filter the dates to only include Bank Holidays
    feriados_nacionais = [date for date in intervalo_datas if date in feriados]
    return feriados_nacionais


def substit(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    subst = session.query(Substituicao).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for s in subst:
        if inicio <= dt.strptime(s.data, '%d/%m/%Y') <= fim:
            if s.professorsubst in dic:
                dic2 = {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                dic[s.professorsubst] = {**dic[s.professorsubst], **dic2[s.professorsubst]}
            else:
                d2 = {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                dic = {**dic, **d2}
    return dic


def desligamentos(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    desl = session.query(Desligados).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for d in desl:
        if inicio <= dt.strptime(d.datadesligamento, '%d/%m/%Y') <= fim:
            if d.professor in dic:
                d2 = {d.professor: {d.departamento: d.datadesligamento}}
                dic[d.professor] = {**dic[d.professor], **d2[d.professor]}
            else:
                d2 = {d.professor: {d.departamento: d.datadesligamento}}
                dic = {**dic, **d2}
    return dic


def escala(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    esc = session.query(Escala).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for e in esc:
        if inicio <= dt.strptime(e.data, '%d/%m/%Y') <= fim:
            if e.professor in dic:
                d2 = {e.professor: {e.data: {e.departamento: e.horas}}}
                dic[e.professor] = {**dic[e.professor], **d2[e.professor]}
            else:
                d2 = {e.professor: {e.data: {e.departamento: e.horas}}}
                dic = {**dic, **d2}
    return dic


def horascomplementares(comp):
    sessions = sessionmaker(bind=engine)
    session = sessions()
    hrsc = session.query(Hrcomplement).all()
    inicio = dt(day=21, month=(dt(day=1, month=comp, year=dt.today().year) - relativedelta(months=1)).month,
                year=dt.today().year)
    fim = dt(day=20, month=comp, year=dt.today().year)
    dic = {}
    for h in hrsc:
        if inicio <= dt.strptime(h.data, '%d/%m/%Y') <= fim:
            if h.professor in dic:
                d2 = {h.professor: {h.data: {h.departamento: h.horas}}}
                dic[h.professor] = {**dic[h.professor], **d2[h.professor]}
            else:
                d2 = {h.professor: {h.data: {h.departamento: h.horas}}}
                dic = {**dic, **d2}
    return dic

atestado = PatternFill(start_color='A9D08E',
                      end_color='A9D08E',
                      fill_type='solid')
falta = PatternFill(start_color='FF0000',
                      end_color='FF0000',
                      fill_type='solid')
ferias = PatternFill(start_color='9BC2E6',
                      end_color='9BC2E6',
                      fill_type='solid')
feriado = PatternFill(start_color='F4B084',
                      end_color='F4B084',
                      fill_type='solid')
fds = PatternFill(start_color='BFBFBF',
                      end_color='BFBFBF',
                      fill_type='solid')
deslig = PatternFill(start_color='454545',
                      end_color='454545',
                      fill_type='solid')
subst = PatternFill(start_color='FFFF00',
                      end_color='FFFF00',
                      fill_type='solid')
comple = PatternFill(start_color='FFC000',
                      end_color='FFC000',
                      fill_type='solid')

def plandegrade(dic, comp):
    grade = l_w('Grade.xlsx', read_only=False)
    plan1 = grade['Planilha1']
    flt = faltas(comp)
    subs = substit(comp)
    dslg = desligamentos(comp)
    fer = feriasf(comp)
    complem = horascomplementares(comp)
    atest = atestadof(comp)
    feriad = feriadof(comp)
    escal = escala(comp)
    competencia = dt(day=10, month=comp, year=dt.today().year)
    inicio = dt(day=21, month=(competencia - relativedelta(months=1)).month, year=(competencia - relativedelta(months=1)).year)
    fechamento = dt(day=20, month=competencia.month, year=competencia.year)
    # primeira linha deve aparecer 'Folha' na coluna A1 e 'Julho' de '2023' na B1
    plan1['A1'].value = 'Folha'
    plan1['B1'].value = f'{fechamento.month} de {fechamento.year}'

    def intervalo(inicio, fechamento):
        for n in range(int((fechamento - inicio).days) + 1):
            yield dt.strftime(inicio + td(n), '%d/%m/%Y')

    col = 3
    for item in list(intervalo(inicio, fechamento)):
        plan1.cell(column=col, row=3, value=dt.strftime(dt.strptime(item, '%d/%m/%Y'), '%a'))
        plan1.cell(column=col, row=4, value=dt.strftime(dt.strptime(item, '%d/%m/%Y'), '%d/%m'))
        col += 1
    plan1.cell(column=col, row=3, value='Total')
    for itens in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
        for cell in itens:
            if cell.value != 'Total':
                if cell.value == 'sáb' or cell.value == 'dom':
                    letras = openpyxl.utils.cell.get_column_letter(cell.column)
                    for numero in range(3,150):
                        plan1[f'{letras}{numero}'].fill = fds
    musculacao = []
    ginastica = []
    esportes = []
    kids = []
    cross = []
    for i in dic:
        for sub in dic[i]:
            if dic[i][sub] == {}:
                pass
            else:
                if sub == 'Musculação':
                    musculacao.append(i)
                    musculacao.sort()
                if sub == 'Ginástica':
                    ginastica.append(i)
                    ginastica.sort()
                if sub == 'Esportes':
                    esportes.append(i)
                    esportes.sort()
                if sub == 'Kids':
                    kids.append(i)
                    kids.sort()
                if sub == 'Cross Cia':
                    cross.append(i)
                    cross.sort()

    plan1['A4'].value = 'Musculação'
    novalinha = 5
    for i in musculacao:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somaseg(i, 'Musculação'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somater(i, 'Musculação'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqua(i, 'Musculação'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqui(i, 'Musculação'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somasex(i, 'Musculação'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somasab(i, 'Musculação'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somadom(i, 'Musculação'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Musculação' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                # {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                # {s.professorsubst: {s.data: {s.substituto: {s.departamento: s.horas}}}}
                for nome in subs:
                    for substituto in subs[nome]:
                        for depart in subs[nome][substituto]:
                            for dia in subs[nome][substituto][depart]:
                                if depart == 'Musculação' and nome == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = falta
                                if depart == 'Musculação' and substituto == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(subs[nome][substituto][depart][dia]).replace(',','.'))
                                        plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                # {d.professor: {d.departamento: d.datadesligamento}}
                # conferir se tem outras aulas ativas ou foi desligado de tudo
                # se desligado de tudo, alterar status das aulas para inativas
                for nome in dslg:
                    for depart in dslg[nome]:
                        for dia in dslg[nome][depart]:
                            if depart == 'Musculação' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= fechamento:
                                        plan1.cell(column=cell.column, row=novalinha).value = 0
                                        plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                        plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                # # {f.professor: {f.departamento: {f.inicio: f.fim}}}
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Musculação' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= dt.strptime(fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                # {h.professor: {h.data: {h.departamento: h.horas}}}
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Musculação' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(complem[nome][dia][depart]).replace(',','.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                # {a.professor: {a.data: a.departamento}}
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Musculação' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                # [datas de feriado formato dt]
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                # {e.professor: {e.data: {e.departamento: e.horas}}}
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Musculação' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(escal[nome][dia][depart]).replace(',','.'))
        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Ginástica'
    novalinha += 1
    for i in ginastica:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somaseg(i, 'Ginástica'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somater(i, 'Ginástica'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqua(i, 'Ginástica'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqui(i, 'Ginástica'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somasex(i, 'Ginástica'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somasab(i, 'Ginástica'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somadom(i, 'Ginástica'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Ginástica' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                # {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                for nome in subs:
                    for substituto in subs[nome]:
                        for depart in subs[nome][substituto]:
                            for dia in subs[nome][substituto][depart]:
                                if depart == 'Ginástica' and nome == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = falta
                                if depart == 'Ginástica' and substituto == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(subs[nome][substituto][depart][dia]).replace(',','.'))
                                        plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                # {d.professor: {d.departamento: d.datadesligamento}}
                # conferir se tem outras aulas ativas ou foi desligado de tudo
                # se desligado de tudo, alterar status das aulas para inativas
                for nome in dslg:
                    for depart in dslg[nome]:
                        for dia in dslg[nome][depart]:
                            if depart == 'Ginástica' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= fechamento:
                                        plan1.cell(column=cell.column, row=novalinha).value = 0
                                        plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                        plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                # # {f.professor: {f.departamento: {f.inicio: f.fim}}}
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Ginástica' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= dt.strptime(fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                # {h.professor: {h.data: {h.departamento: h.horas}}}
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Ginástica' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(complem[nome][dia][depart]).replace(',','.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                # {a.professor: {a.data: a.departamento}}
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Ginástica' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                # [datas de feriado formato dt]
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                # {e.professor: {e.data: {e.departamento: e.horas}}}
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Ginástica' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(escal[nome][dia][depart]).replace(',','.'))

        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Kids'
    novalinha += 1
    for i in kids:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somaseg(i, 'Kids'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somater(i, 'Kids'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqua(i, 'Kids'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqui(i, 'Kids'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somasex(i, 'Kids'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somasab(i, 'Kids'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somadom(i, 'Kids'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Kids' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                # {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                for nome in subs:
                    for substituto in subs[nome]:
                        for depart in subs[nome][substituto]:
                            for dia in subs[nome][substituto][depart]:
                                if depart == 'Kids' and nome == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = falta
                                if depart == 'Kids' and substituto == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(subs[nome][substituto][depart][dia]).replace(',','.'))
                                        plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                # {d.professor: {d.departamento: d.datadesligamento}}
                # conferir se tem outras aulas ativas ou foi desligado de tudo
                # se desligado de tudo, alterar status das aulas para inativas
                for nome in dslg:
                    for depart in dslg[nome]:
                        if depart == 'Kids' and nome == i and dt.strptime(dslg[nome][depart], '%d/%m/%Y') <= fechamento:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                if dt.strptime(dslg[nome][depart], '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= fechamento:
                                    plan1.cell(column=cell.column, row=novalinha).value = 0
                                    plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                    plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                # # {f.professor: {f.departamento: {f.inicio: f.fim}}}
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Kids' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= dt.strptime(fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                # {h.professor: {h.data: {h.departamento: h.horas}}}
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Kids' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(complem[nome][dia][depart]).replace(',','.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                # {a.professor: {a.data: a.departamento}}
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Kids' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                # [datas de feriado formato dt]
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                # {e.professor: {e.data: {e.departamento: e.horas}}}
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Kids' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(escal[nome][dia][depart]).replace(',','.'))
        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Esportes'
    novalinha += 1
    for i in esportes:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somaseg(i, 'Esportes'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somater(i, 'Esportes'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqua(i, 'Esportes'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqui(i, 'Esportes'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somasex(i, 'Esportes'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somasab(i, 'Esportes'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somadom(i, 'Esportes'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Esportes' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                # {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                for nome in subs:
                    for substituto in subs[nome]:
                        for depart in subs[nome][substituto]:
                            for dia in subs[nome][substituto][depart]:
                                if depart == 'Esportes' and nome == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = falta
                                if depart == 'Esportes' and substituto == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(subs[nome][substituto][depart][dia]).replace(',','.'))
                                        plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                # {d.professor: {d.departamento: d.datadesligamento}}
                # conferir se tem outras aulas ativas ou foi desligado de tudo
                # se desligado de tudo, alterar status das aulas para inativas
                for nome in dslg:
                    for depart in dslg[nome]:
                        for dia in dslg[nome][depart]:
                            if depart == 'Esportes' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= fechamento:
                                        plan1.cell(column=cell.column, row=novalinha).value = 0
                                        plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                        plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                # {f.professor: {f.departamento: {f.inicio: f.fim}}}
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Esportes' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= dt.strptime(fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                # {h.professor: {h.data: {h.departamento: h.horas}}}
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Esportes' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(complem[nome][dia][depart]).replace(',','.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                # {a.professor: {a.data: a.departamento}}
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Esportes' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                # [datas de feriado formato dt]
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                # {e.professor: {e.data: {e.departamento: e.horas}}}
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Esportes' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(escal[nome][dia][depart]).replace(',','.'))
        
        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    plan1[f'A{novalinha}'].value = 'Cross Cia'
    novalinha += 1
    for i in cross:
        for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
            for cell in row:
                if cell.value != 'Total':
                    if cell.value == 'seg':
                        plan1.cell(column=cell.column, row=novalinha, value=somaseg(i, 'Cross Cia'))
                    if cell.value == 'ter':
                        plan1.cell(column=cell.column, row=novalinha, value=somater(i, 'Cross Cia'))
                    if cell.value == 'qua':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqua(i, 'Cross Cia'))
                    if cell.value == 'qui':
                        plan1.cell(column=cell.column, row=novalinha, value=somaqui(i, 'Cross Cia'))
                    if cell.value == 'sex':
                        plan1.cell(column=cell.column, row=novalinha, value=somasex(i, 'Cross Cia'))
                    if cell.value == 'sáb':
                        plan1.cell(column=cell.column, row=novalinha, value=somasab(i, 'Cross Cia'))
                    if cell.value == 'dom':
                        plan1.cell(column=cell.column, row=novalinha, value=somadom(i, 'Cross Cia'))
                else:
                    letra = openpyxl.utils.cell.get_column_letter(cell.column - 1)
                    plan1.cell(column=cell.column, row=novalinha, value=f'=SUM(C{novalinha}:{letra}{novalinha})')
                # aplica cor de falta
                for nome in flt:
                    for dia in flt[nome]:
                        for depart in flt[nome][dia]:
                            if depart == 'Cross Cia' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).fill = falta
                # aplica alterações de substituição
                # {s.professorsubst: {s.substituto: {s.departamento: {s.data: s.horas}}}}
                for nome in subs:
                    for substituto in subs[nome]:
                        for depart in subs[nome][substituto]:
                            for dia in subs[nome][substituto][depart]:
                                if depart == 'Cross Cia' and nome == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = falta
                                if depart == 'Cross Cia' and substituto == i:
                                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                        plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(subs[nome][substituto][depart][dia]).replace(',','.'))
                                        plan1.cell(column=cell.column, row=novalinha).fill = subst

                # aplica alterações de desligamento
                # {d.professor: {d.departamento: d.datadesligamento}}
                # conferir se tem outras aulas ativas ou foi desligado de tudo
                # se desligado de tudo, alterar status das aulas para inativas
                for nome in dslg:
                    for depart in dslg[nome]:
                        for dia in dslg[nome][depart]:
                            if depart == 'Cross Cia' and nome == i and dt.strptime(dia, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(dia, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= fechamento:
                                        plan1.cell(column=cell.column, row=novalinha).value = 0
                                        plan1.cell(column=cell.column, row=novalinha).fill = deslig
                                        plan1.cell(column=cell.column, row=novalinha).font = Font(color='FFFFFF')

                # aplica talterações de férias
                # # {f.professor: {f.departamento: {f.inicio: f.fim}}}
                for nome in fer:
                    for depart in fer[nome]:
                        for inic in fer[nome][depart]:
                            if depart == 'Cross Cia' and nome == i and dt.strptime(inic, '%d/%m/%Y') <= fechamento:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value is not None:
                                    if dt.strptime(inic, '%d/%m/%Y') <= dt(day=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[0]), month=int(str(plan1.cell(column=cell.column, row=cell.row + 1).value).split('/')[1]), year=dt.today().year) <= dt.strptime(fer[nome][depart][inic], '%d/%m/%Y'):
                                        plan1.cell(column=cell.column, row=novalinha).fill = ferias
                                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de horas complementares
                # {h.professor: {h.data: {h.departamento: h.horas}}}
                for nome in complem:
                    for dia in complem[nome]:
                        for depart in complem[nome][dia]:
                            if depart == 'Cross Cia' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(complem[nome][dia][depart]).replace(',','.'))
                                    plan1.cell(column=cell.column, row=novalinha).fill = comple

                # aplica alterações de atestados
                # {a.professor: {a.data: a.departamento}}
                for nome in atest:
                    for d in atest[nome]:
                        if atest[nome][d] == 'Cross Cia' and nome == i:
                            if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(d, '%d/%m/%Y'), '%d/%m'):
                                plan1.cell(column=cell.column, row=novalinha).fill = atestado

                # aplica alterações de feriado
                # [datas de feriado formato dt]
                for dia in feriad:
                    if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dia, '%d/%m'):
                        plan1.cell(column=cell.column, row=novalinha).fill = feriado
                        plan1.cell(column=cell.column, row=novalinha).value = 0

                # aplica alterações de escala
                # {e.professor: {e.data: {e.departamento: e.horas}}}
                for nome in escal:
                    for dia in escal[nome]:
                        for depart in escal[nome][dia]:
                            if depart == 'Cross Cia' and nome == i:
                                if plan1.cell(column=cell.column, row=cell.row + 1).value == dt.strftime(dt.strptime(dia, '%d/%m/%Y'), '%d/%m'):
                                    plan1.cell(column=cell.column, row=novalinha).value = plan1.cell(column=cell.column, row=novalinha).value + float(str(escal[nome][dia][depart]).replace(',','.'))

        plan1.cell(column=2, row=novalinha, value=i)
        novalinha += 1

    for i, coluna in enumerate(plan1.columns):
        max_length = 0
        column = coluna[0].column_letter
        for cell in coluna:
            try:
                if len(str(cell.value)) > max_length and len(str(cell.value)) != 18:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 1
        plan1.column_dimensions[column].width = adjusted_width

    for row in plan1.iter_cols(min_row=3, min_col=3, max_row=3, max_col=35):
        for cell in row:
            if plan1.cell(column=cell.column, row=cell.row).value == 'Total':
                plan1.column_dimensions[openpyxl.utils.cell.get_column_letter(cell.column)].width = 8

    plan1['C1'].fill = atestado
    plan1['D1'].value = 'Atestado'
    plan1['C2'].fill = falta
    plan1['D2'].value = 'Falta'
    plan1['F1'].fill = ferias
    plan1['G1'].value = 'Férias'
    plan1['F2'].fill = feriado
    plan1['G2'].value = 'Feriado'
    plan1['I1'].fill = deslig
    plan1['J1'].value = 'Desligamento'
    plan1['I2'].fill = subst
    plan1['J2'].value = 'Substituiu'
    plan1['M1'].fill = comple
    plan1['N1'].value = 'Horas Complementares'
    # plan1[]
    grade.save(f'Grade {fechamento.month}-{fechamento.year}.xlsx')


def somaseg(nome, depto):
    aulasseg = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Segunda')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasseg:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def somater(nome, depto):
    aulaster = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Terça')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulaster:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def somaqua(nome, depto):
    aulasqua = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Quarta')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasqua:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def somaqui(nome, depto):
    aulasqui = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Quinta')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasqui:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def somasex(nome, depto):
    aulassex = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Sexta')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulassex:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def somasab(nome, depto):
    aulassab = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Sábado')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulassab:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def somadom(nome, depto):
    aulasdom = session.query(Aulas).filter_by(professor=nome).filter_by(diadasemana='Domingo')\
        .filter_by(departamento=depto).all()
    somas = 0
    for aula in aulasdom:
        hr, minut, seg = str(dt.strptime(aula.fim, '%H:%M:%S') - dt.strptime(aula.inicio, '%H:%M:%S')).split(':')
        somadia = int(hr) + int(minut) / 60 + int(seg) / 60 * 60
        somadia = round(somadia, 2)
        somas += somadia
    return somas


def plansoma():
    folhadehoje = Folha(data, list(aulasativas()), deptosativos())
    somaaulas = {}
    for i in profsativos():
        somaaulas[i] = {}
        for d in deptosativos():
            somaaulas[i][d] = {}
    for aulas in aulasativas():
        somaaulas[aulas.professor][aulas.departamento][aulas.nome+f' ({aulas.valor})'] = round(somaprof(folhadehoje, aulas.professor, aulas.departamento, aulas.nome), 2)
        dictchav = list(somaaulas.keys())
        dictchav.sort()
        somafinal = {i: somaaulas[i] for i in dictchav}

    plan = l_w('Somafinal.xlsx', read_only=False)
    folha = plan['Planilha1']
    folha['A1'].value = 'Matrícula'
    folha['B1'].value = 'Nome'
    folha['C1'].value = 'Aula e Valor'
    folha['D1'].value = 'Horas'
    x = 2
    for i in somafinal:
        matr = session.query(Aulas).filter_by(professor=str(i)).first()
        folha[f'A{x}'].value = int(matr.matrprof)
        folha[f'B{x}'].value = str(i)
        for sub in somafinal[i]:
            if somafinal[i][sub] == {}:
                pass
            else:
                for sub2 in somafinal[i][sub]:
                    folha[f'C{x}'].value = str(sub2)
                    folha[f'D{x}'].value = float(str(somafinal[i][sub][sub2]))
                    x += 1
    # folha['F1'].value = 'Total Bruto - Professores'
    # folha['G1'].value = locale.currency(totaldafolha(folhadehoje), grouping=True)
    plan.save(f'Somafinal mes {data}.xlsx')
    plandegrade(somafinal, data)
    substitutos = {}
    complementares = {}
    feriasl = {}
    desligadosl = {}
    planilha = l_w(f'Grade {data}-2023.xlsx')
    aba = planilha['Planilha1']
    for row in aba.iter_cols(min_row=3, min_col=3, max_row=115, max_col=35):
        for cell in row:
            if cell.fill == ferias:
                for i in range(1,150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                feriasl[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}
            if cell.fill == deslig:
                for i in range(1,150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                desligadosl[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}
            if cell.fill == subst:
                for i in range(1,150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                substitutos[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}
            if cell.fill == comple:
                for i in range(1, 150):
                    if aba.cell(column=1, row=cell.row - i).value is not None:
                        depart = aba.cell(column=1, row=cell.row - i).value
                        break
                for r in aba.iter_cols(min_row=3, min_col=3, max_row=3, max_col=38):
                    for c in r:
                        if c.value == 'Total':
                            tt = c.column
                hrs = 0
                for m in range(3, tt):
                    hrs += aba.cell(column=m, row=cell.row).value
                complementares[aba.cell(column=2, row=cell.row).value] = {depart: round(hrs, 2)}

    planilha2 = l_w(f'Somafinal mes {data}.xlsx', read_only=False)
    aba2 = planilha2['Planilha1']
    for pessoa in feriasl:
        for depart in feriasl[pessoa]:
            for cll in aba2['B']:
                if cll.value is not None:
                    if cll.value == pessoa:
                        aba2.cell(column=4, row=cll.row, value=float(feriasl[pessoa][depart]))
    for pessoa in desligadosl:
        for depart in desligadosl[pessoa]:
            for cell in aba2['B']:
                if cell.value is not None:
                    if cell.value == pessoa:
                        aba2.cell(column=4, row=cell.row).value = float(desligadosl[pessoa][depart]) 
    for pessoa in substitutos:
        for depart in substitutos[pessoa]:
            for cell in aba2['B']:
                if cell.value is not None:
                    if cell.value == pessoa:
                        aba2.cell(column=4, row=cell.row).value = float(substitutos[pessoa][depart]) 
    for pessoa in complementares:
        for depart in complementares[pessoa]:
            for cell in aba2['B']:
                if cell.value is not None:
                    if cell.value == pessoa:
                        aba2.cell(column=4, row=cell.row).value = float(complementares[pessoa][depart]) 
    for i, coluna in enumerate(aba2.columns):
        max_length = 0
        column = coluna[0].column_letter
        for cell in coluna:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 1
        aba2.column_dimensions[column].width = adjusted_width
    planilha2.save(f'Somafinal mes {data}.xlsx')
    print('Férias \n', feriasl)
    print('Desligados \n', desligadosl)
    print('Substitutos \n', substitutos)
    print('Hrs Complementares \n', complementares)


plansoma()
