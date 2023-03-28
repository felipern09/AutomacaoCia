import pandas as pd
from datetime import datetime, timedelta
import locale
from openpyxl import load_workbook

locale.setlocale(locale.LC_ALL, 'pt_pt.UTF-8')
data_inicial = datetime.strptime('01/10/2022', '%d/%m/%Y')
data_final = datetime.strptime('13/11/2022', '%d/%m/%Y')

# Ler planilha geral
geral = pd.read_excel('../Ponto/xls/zzPonto Geral.xls')
geral = geral.rename(
    columns={'CARTÃO PONTO': 'Dia', 'Unnamed: 1': 0, 'Unnamed: 2': 1, 'Unnamed: 3': 2, 'Unnamed: 4': 3, 'Unnamed: 5': 4,
             'Unnamed: 6': 5})
geral = geral.drop(['Unnamed: 7', 'Unnamed: 8'], axis=1)
geral = geral[geral.Dia.notnull()]

# Pegar index onde aparece 'Nome'
linhasNomes = geral.index[geral['Dia'].str.contains('Nome')]

# salvar plan com nome do funcionário pasta ponto (dentro da pasta automação)
for linha in linhasNomes:
    geral = geral.rename(
        columns={'CARTÃO PONTO': 'Dia', 'Unnamed: 1': 0, 'Unnamed: 2': 1, 'Unnamed: 3': 2, 'Unnamed: 4': 3,
                 'Unnamed: 5': 4, 'Unnamed: 6': 5})
    geral = geral[geral.Dia.notnull()]
    geral = geral[geral['Dia'].str.contains(' - ') | geral['Dia'].str.contains('Nome')]
    geral2 = geral.loc[linha:(linha + (linhasNomes[1] - linhasNomes[0] - 1))]\
        .to_excel(f'../Ponto/xls/{geral[0][linha]}.xlsx')

    # Verifica a hora certa na planilha zzBase.xlsx
    wb = load_workbook(f'../Ponto/xls/zzBase.xlsx')
    ws = wb.active
    for row in ws.rows:
        for cell in row:
            if cell.value == f'{geral[0][linha]}':
                if ws.cell(row=cell.row, column=5).value is None:
                    ent1 = datetime.strptime('00:00:00', '%H:%M:%S')
                else:
                    ent1 = datetime.strptime(str(ws.cell(row=cell.row, column=5).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=6).value is None:
                    sai1 = datetime.strptime('00:00:00', '%H:%M:%S')
                else:
                    sai1 = datetime.strptime(str(ws.cell(row=cell.row, column=6).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=7).value is None:
                    ent2 = datetime.strptime('00:00:00', '%H:%M:%S')
                else:
                    ent2 = datetime.strptime(str(ws.cell(row=cell.row, column=7).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=8).value is None:
                    sai2 = datetime.strptime('00:00:00', '%H:%M:%S')
                else:
                    sai2 = datetime.strptime(str(ws.cell(row=cell.row, column=8).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=9).value is None:
                    ent3 = datetime.strptime('00:00:00', '%H:%M:%S')
                else:
                    ent3 = datetime.strptime(str(ws.cell(row=cell.row, column=9).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=10).value is None:
                    sai3 = datetime.strptime('00:00:00', '%H:%M:%S')
                else:
                    sai3 = datetime.strptime(str(ws.cell(row=cell.row, column=10).value), '%H:%M:%S')
    # relacionar dia da semana
    entradacerta = timedelta(hours=ent1.hour, minutes=ent1.minute, seconds=ent1.second)
    saidacerta = timedelta(hours=sai1.hour, minutes=sai1.minute, seconds=sai1.second)
    entradacerta2 = timedelta(hours=ent2.hour, minutes=ent2.minute, seconds=ent2.second)
    saidacerta2 = timedelta(hours=sai2.hour, minutes=sai2.minute, seconds=sai2.second)
    entradacerta3 = timedelta(hours=ent3.hour, minutes=ent3.minute, seconds=ent3.second)
    saidacerta3 = timedelta(hours=sai3.hour, minutes=sai3.minute, seconds=sai3.second)
    regra = timedelta(hours=0, minutes=10, seconds=0)

    # Salvar planilha adicionando colunas de diferenças
    plan = pd.read_excel(f'../Ponto/xls/{geral[0][linha]}.xlsx')
    plan = plan[plan['Dia'].str.contains(' - ')]
    plan = plan.drop(['Unnamed: 0'], axis=1)
    plan = plan.rename(columns={'Dia': 'Data'})
    plan[0] = pd.to_timedelta(plan[0].astype(str))
    plan[1] = pd.to_timedelta(plan[1].astype(str))
    plan[2] = pd.to_timedelta(plan[2].astype(str))
    plan[3] = pd.to_timedelta(plan[3].astype(str))
    plan[4] = pd.to_timedelta(plan[4].astype(str))
    plan[5] = pd.to_timedelta(plan[5].astype(str))
    plan['Entradacerta'] = entradacerta
    plan['Saídacerta'] = saidacerta
    plan['DifEntr'] = abs(plan[0] - plan['Entradacerta'])
    plan['DifSaida'] = abs(plan[1] - plan['Saídacerta'])
    dif = plan
    dif = dif[['Data', 'DifEntr', 'DifSaida']]
    dif = dif.loc[(dif['DifEntr'] >= regra) | (dif['DifEntr'] + dif['DifSaida'] >= regra)]
    dif = dif.astype(str)
    dif['DifEntr'] = dif['DifEntr'].map(
        lambda x: datetime
        .strftime(datetime.strptime(str(x).replace('0 days ', '').replace('NaT', '00:00:00'), '%H:%M:%S'),
                  '%H hora(s) e %M minutos')
    )
    dif['DifSaida'] = dif['DifSaida'].map(
        lambda x: datetime.strftime(datetime.strptime(str(x).replace('0 days ', '').replace('NaT', '00:00:00'),
                                                      '%H:%M:%S'), '%H hora(s) e %M minutos')
    )
    dif['DifEntr'] = dif['DifEntr'].map(
        lambda x: str(x).replace('00 hora(s) e 00 minutos', '-').replace('00 hora(s) e ', '')
    )
    dif['DifSaida'] = dif['DifSaida'].map(
        lambda x: str(x).replace('00 hora(s) e 00 minutos', '-').replace('00 hora(s) e ', '')
    )
    plan = plan[['Data', 0, 1, 2, 3]]
    plan = plan.merge(dif, on='Data', how='outer')
    plan = plan.astype(str)
    # plan = plan.fillna('-')
    # plan['Data'] = plan['Data'].map(lambda x: x.rstrip('- qua qui sex sab ter seg sá dom'))
    plan[0] = plan[0].map(lambda x: x.lstrip('0 days 00:'))
    plan[1] = plan[1].map(lambda x: x.lstrip('0 days 00:'))
    plan[2] = plan[2].map(lambda x: x.lstrip('0 days 00:'))
    plan[3] = plan[3].map(lambda x: x.lstrip('0 days 00:'))
    plan = plan.replace('NaT', '-').replace('nan', '-')
    plan = plan.rename(columns={'Data': 'Dia', 0: 'Entrada1', 1: 'Saída 1', 2: 'Entrada 2', 3: 'Saída 2', 'DifEntr':
                                'Diferença Entrada', 'DifSaida': 'Diferença Saída 1'})
    plan = plan[plan.Entrada1 != '-']
    plan = plan.rename(columns={'Entrada1': 'Entrada 1'})
    plan = plan.astype(str)
    plan['Dia'] = plan['Dia'].map(lambda x: datetime.strptime(x, '%d/%m/%y - %a'))
    plan = plan[plan.Dia >= data_inicial]
    plan = plan[plan.Dia <= data_final]
    plan['Dia'] = plan['Dia'].map(lambda x: datetime.strftime(x, '%d/%m/%y - %a'))
    plan = plan.rename(columns={'Dia': 'Data'})
    plan = plan.set_index(['Data'])
    plan = plan.to_excel(f'../Ponto/xls/{geral[0][linha]}.xlsx')
    func = load_workbook(f'../Ponto/xls/{geral[0][linha]}.xlsx', read_only=False)
    sh = func['Sheet1']
    sh.column_dimensions['A'].width = 15
    sh.column_dimensions['F'].width = 22
    sh.column_dimensions['G'].width = 22
    sh.column_dimensions['H'].width = 22
    sh.column_dimensions['I'].width = 22
    sh.column_dimensions['J'].width = 22
    sh.column_dimensions['K'].width = 22
    func.save(f'../Ponto/xls/{geral[0][linha]}.xlsx')
