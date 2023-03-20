from openpyxl import load_workbook as l_w
from openpyxl.styles import NamedStyle
import pandas as pd
from datetime import datetime as dt
from datetime import timedelta
import os

# Ler arquivo txt dos registror rejeitados
geral = pd.read_csv('Rejeitados.txt', sep=' ', header=None, encoding='iso8859-1')
geral = geral.rename(columns={0: 'matricula',16: 'data', 17: 'dia', 18: 'hora'})
geral = geral[geral.dia != 'Batida']
geral = geral[geral.matricula < 9999]
mat =[]
mat_unicas=[]
for matricula in geral['matricula']:
    mat.append(matricula)
    mat_unicas= list(set(mat))
for matr in mat_unicas:
    geral2 = geral[geral.matricula == matr]
    geral2 = geral2.set_index('matricula')
    geral2 = geral2.dropna(axis='columns')
    geral2 = geral2.drop(geral2.iloc[:, [2,3,4,5]], axis=1)
    geral2['dia'] = pd.to_datetime(geral2['dia'], format='%d/%m/%Y')
    geral2['dia'] = geral2['dia'].apply(lambda y: dt.strftime(y, '%d/%m/%Y'))
    geral2['hora'] = geral2['hora'].apply(lambda x: f'{x}:00')
    geral2['hora'] = pd.to_timedelta(geral2['hora'])
    # geral2 = geral2.set_index('dia')
    # geral2 = geral2.groupby('dia')
    geral2.to_excel(f'../Ponto/xls/Ponto Estágio - {matr}.xlsx')
    wb = l_w(f'../Ponto/xls/Ponto Estágio - {matr}.xlsx')
    sh = wb['Sheet1']
    sh['A1'].value = 'Matrícula'
    sh['B1'].value = 'Data'
    sh['C1'].value = 'Entrada 1'
    sh['D1'].value = 'Saída 1'
    sh['E1'].value = 'Entrada 2'
    sh['F1'].value = 'Saída 2'
    sh['G1'].value = 'Entrada 3'
    sh['H1'].value = 'Saída 3'
    x = 2
    for row in sh:
        if sh[f'B{x}'].value == sh[f'B{x-1}'].value:
            if sh[f'D{x-1}'].value is None:
                sh[f'D{x - 1}'].value = sh[f'C{x}'].value
                sh.delete_rows(x, 1)
            else:
                if sh[f'E{x - 1}'].value is None:
                    sh[f'E{x - 1}'].value = sh[f'C{x}'].value
                    sh.delete_rows(x, 1)
                else:
                    if sh[f'F{x - 1}'].value is None:
                        sh[f'F{x - 1}'].value = sh[f'C{x}'].value
                        sh.delete_rows(x, 1)
                    else:
                        if sh[f'G{x - 1}'].value is None:
                            sh[f'G{x - 1}'].value = sh[f'C{x}'].value
                            sh.delete_rows(x, 1)
                        else:
                            if sh[f'H{x - 1}'].value is None:
                                sh[f'H{x - 1}'].value = sh[f'C{x}'].value
                                sh.delete_rows(x, 1)
    for row in sh:
        if sh[f'B{x}'].value == sh[f'B{x - 1}'].value:
            if sh[f'D{x - 1}'].value is None:
                sh[f'D{x - 1}'].value = sh[f'C{x}'].value
                sh.delete_rows(x, 1)
            else:
                if sh[f'E{x - 1}'].value is None:
                    sh[f'E{x - 1}'].value = sh[f'C{x}'].value
                    sh.delete_rows(x, 1)
                else:
                    if sh[f'F{x - 1}'].value is None:
                        sh[f'F{x - 1}'].value = sh[f'C{x}'].value
                        sh.delete_rows(x, 1)
                    else:
                        if sh[f'G{x - 1}'].value is None:
                            sh[f'G{x - 1}'].value = sh[f'C{x}'].value
                            sh.delete_rows(x, 1)
                        else:
                            if sh[f'H{x - 1}'].value is None:
                                sh[f'H{x - 1}'].value = sh[f'C{x}'].value
                                sh.delete_rows(x, 1)
        for row in sh:
            if sh[f'B{x}'].value == sh[f'B{x - 1}'].value:
                if sh[f'D{x - 1}'].value is None:
                    sh[f'D{x - 1}'].value = sh[f'C{x}'].value
                    sh.delete_rows(x, 1)
                else:
                    if sh[f'E{x - 1}'].value is None:
                        sh[f'E{x - 1}'].value = sh[f'C{x}'].value
                        sh.delete_rows(x, 1)
                    else:
                        if sh[f'F{x - 1}'].value is None:
                            sh[f'F{x - 1}'].value = sh[f'C{x}'].value
                            sh.delete_rows(x, 1)
                        else:
                            if sh[f'G{x - 1}'].value is None:
                                sh[f'G{x - 1}'].value = sh[f'C{x}'].value
                                sh.delete_rows(x, 1)
                            else:
                                if sh[f'H{x - 1}'].value is None:
                                    sh[f'H{x - 1}'].value = sh[f'C{x}'].value
                                    sh.delete_rows(x, 1)
        x += 1

    estilo_data = NamedStyle(name='data', number_format='DD/MM/YYYY')
    estilo_hora = NamedStyle(name='hora', number_format='HH:MM:SS')
    for cell in sh['B']:
        sh[f'B{int(cell.row)}'].style = estilo_data
    for item in sh['C']:
        sh[f'C{int(item.row)}'].style = estilo_hora
    for item in sh['D']:
        sh[f'D{int(item.row)}'].style = estilo_hora
    for item in sh['E']:
        sh[f'E{int(item.row)}'].style = estilo_hora
    for item in sh['F']:
        sh[f'F{int(item.row)}'].style = estilo_hora
    for item in sh['G']:
        sh[f'G{int(item.row)}'].style = estilo_hora
    for item in sh['H']:
        sh[f'H{int(item.row)}'].style = estilo_hora
    sh.column_dimensions['A'].width = 11
    sh.column_dimensions['B'].width = 11
    sh.column_dimensions['C'].width = 11
    sh.column_dimensions['D'].width = 11
    sh.column_dimensions['E'].width = 11
    sh.column_dimensions['F'].width = 11
    sh.column_dimensions['G'].width = 11
    sh.column_dimensions['H'].width = 11
    wb2 = l_w(f'../Ponto/xls/zzBase.xlsx')
    ws2 = wb2.active
    for row in ws2.rows:
        for cell in row:
            if cell.value == matr:
                nome = ws2.cell(row=cell.row, column=3).value
                wb.save(f'../Ponto/xls/Ponto Estágio - {nome}.xlsx')
                os.remove(f'../Ponto/xls/Ponto Estágio - {matr}.xlsx')
    plan = l_w(f'../Ponto/xls/Ponto Estágio - {nome}.xlsx')
    splan = plan.active
    # Verifica a hora certa na planilha zzBase.xlsx
    wb = l_w(f'../Ponto/xls/zzBase.xlsx')
    ws = wb.active
    for row in ws.rows:
        for cell in row:
            if cell.value == f'{nome}':
                for rowsplan in splan.rows:
                    for cellplan in rowsplan:
                        if str(splan.cell(row=cellplan.row, column=2).value) == 'Data':
                            pass
                        # se o dia no ponto for segunda
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 1:
                            pass
                        # se o dia no ponto for terça
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 2:
                            pass
                        # se o dia no ponto for quarta
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 3:
                            pass
                        # se o dia no ponto for quinta
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 4:
                            pass
                        # se o dia no ponto for sexta
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 5:
                            pass
                        # se o dia no ponto for sábado
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 6:
                            pass
                        # se o dia no ponto for domingo
                        if dt.strptime(str(splan.cell(row=cellplan.row, column=2).value), '%d/%m/%Y').isoweekday() == 7:
                            pass
                if ws.cell(row=cell.row, column=5).value is None:
                    ent1 = dt.strptime('00:00:00', '%H:%M:%S')
                else:
                    ent1 = dt.strptime(str(ws.cell(row=cell.row, column=5).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=6).value is None:
                    sai1 = dt.strptime('00:00:00', '%H:%M:%S')
                else:
                    sai1 = dt.strptime(str(ws.cell(row=cell.row, column=6).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=7).value is None:
                    ent2 = dt.strptime('00:00:00', '%H:%M:%S')
                else:
                    ent2 = dt.strptime(str(ws.cell(row=cell.row, column=7).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=8).value is None:
                    sai2 = dt.strptime('00:00:00', '%H:%M:%S')
                else:
                    sai2 = dt.strptime(str(ws.cell(row=cell.row, column=8).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=9).value is None:
                    ent3 = dt.strptime('00:00:00', '%H:%M:%S')
                else:
                    ent3 = dt.strptime(str(ws.cell(row=cell.row, column=9).value), '%H:%M:%S')
                if ws.cell(row=cell.row, column=10).value is None:
                    sai3 = dt.strptime('00:00:00', '%H:%M:%S')
                else:
                    sai3 = dt.strptime(str(ws.cell(row=cell.row, column=10).value), '%H:%M:%S')
            # relacionar dia da semana
            entradacerta = timedelta(hours=ent1.hour, minutes=ent1.minute, seconds=ent1.second)
            saidacerta = timedelta(hours=sai1.hour, minutes=sai1.minute, seconds=sai1.second)
            entradacerta2 = timedelta(hours=ent2.hour, minutes=ent2.minute, seconds=ent2.second)
            saidacerta2 = timedelta(hours=sai2.hour, minutes=sai2.minute, seconds=sai2.second)
            entradacerta3 = timedelta(hours=ent3.hour, minutes=ent3.minute, seconds=ent3.second)
            saidacerta3 = timedelta(hours=sai3.hour, minutes=sai3.minute, seconds=sai3.second)
            regra = timedelta(hours=0, minutes=10, seconds=0)

# # procurar na planilha base.xlsx nome e-mail matricula no ponto e horarios de entrada e saida




